"""
jira_extractor.py
-----------------
Production-ready Jira → Excel extractor.

Layout
    Column A  : Epic (key + name), merged vertically across all its stories
    Columns B+ : One row per Story — fixed columns + any custom_fields from config

Usage
    python jira_extractor.py                      # uses config.json in CWD
    python jira_extractor.py --config /path/to/config.json

Dependencies
    pip install requests openpyxl msal pywin32
    (pywin32 is required for Windows broker/SSO support)
"""

import argparse
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

import requests
from requests.auth import HTTPBasicAuth
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

import msal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# Config loader
# ──────────────────────────────────────────────────────────────────────────────

VALID_TYPES = {"text", "date", "number", "user", "array"}

def load_config(path: str) -> dict:
    cfg_path = Path(path)
    if not cfg_path.exists():
        raise FileNotFoundError(f"Config file not found: {cfg_path.resolve()}")
    with cfg_path.open("r", encoding="utf-8") as f:
        cfg = json.load(f)

    # Required keys
    for key in ("jira.url", "jira.email", "jira.token",
                "project.key", "project.output"):
        section, field = key.split(".")
        if not cfg.get(section, {}).get(field):
            raise ValueError(f"Missing required config key: {key}")

    # Validate custom_fields entries
    for cf in cfg.get("custom_fields", []):
        if not cf.get("id"):
            raise ValueError(f"custom_fields entry missing 'id': {cf}")
        if not cf.get("label"):
            raise ValueError(f"custom_fields entry missing 'label': {cf}")
        cf_type = cf.get("type", "text")
        if cf_type not in VALID_TYPES:
            raise ValueError(
                f"custom_fields '{cf['id']}' has invalid type '{cf_type}'. "
                f"Must be one of: {', '.join(VALID_TYPES)}"
            )

    # Validate SharePoint config if present
    sp = cfg.get("sharepoint", {})
    if sp.get("enabled", False):
        for key in ("site_url", "library", "folder", "client_id", "tenant_id"):
            if not sp.get(key):
                raise ValueError(f"sharepoint.{key} is required when sharepoint.enabled=true")

    return cfg


# ──────────────────────────────────────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────────────────────────────────────

def setup_logging(cfg: dict) -> logging.Logger:
    log_cfg  = cfg.get("logging", {})
    level    = getattr(logging, log_cfg.get("level", "INFO").upper(), logging.INFO)
    log_file = log_cfg.get("file", "jira_extractor.log")

    handlers = [logging.StreamHandler(sys.stdout)]
    if log_file:
        handlers.append(logging.FileHandler(log_file, encoding="utf-8"))

    logging.basicConfig(
        level=level,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=handlers,
    )
    return logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Jira API client
# ──────────────────────────────────────────────────────────────────────────────

class JiraClient:
    def __init__(self, cfg: dict, logger: logging.Logger):
        jira_cfg         = cfg["jira"]
        fetch_cfg        = cfg.get("fetch", {})
        self.base        = jira_cfg["url"].rstrip("/")
        self.auth        = HTTPBasicAuth(jira_cfg["email"], jira_cfg["token"])
        self.max_results = int(fetch_cfg.get("max_results", 100))
        self.timeout     = int(fetch_cfg.get("timeout_secs", 30))
        self.logger      = logger
        self.session     = self._build_session()

    def _build_session(self) -> requests.Session:
        session = requests.Session()
        session.headers.update({
            "Accept":       "application/json",
            "Content-Type": "application/json",
        })
        session.auth = self.auth
        retry = Retry(
            total=3,
            backoff_factor=1,
            status_forcelist=[429, 500, 502, 503, 504],
            allowed_methods=["GET", "POST"],
        )
        session.mount("https://", HTTPAdapter(max_retries=retry))
        session.mount("http://",  HTTPAdapter(max_retries=retry))
        return session

    def fetch_all(self, jql: str, fields: list[str]) -> list[dict]:
        """
        Paginate POST /rest/api/3/search/jql and return all matching issues.
        Uses cursor-based pagination via nextPageToken.
        """
        issues          = []
        next_page_token = None
        self.logger.debug("JQL: %s | fields: %s", jql, fields)

        while True:
            payload: dict = {
                "jql":        jql,
                "fields":     fields,
                "maxResults": self.max_results,
            }
            if next_page_token:
                payload["nextPageToken"] = next_page_token

            resp = self.session.post(
                f"{self.base}/rest/api/3/search/jql",
                json=payload,
                timeout=self.timeout,
            )
            resp.raise_for_status()
            data  = resp.json()
            batch = data.get("issues", [])
            issues.extend(batch)
            self.logger.debug("Fetched %d issues so far", len(issues))

            next_page_token = data.get("nextPageToken")
            if not next_page_token or not batch:
                break

        return issues


# ──────────────────────────────────────────────────────────────────────────────
# SharePoint client  (SharePoint Online — Windows SSO, no stored credentials)
# ──────────────────────────────────────────────────────────────────────────────

class SharePointClient:
    """
    Uploads/overwrites a file in SharePoint Online using your existing
    Windows AD login session — no username or password stored anywhere.

    Authentication flow:
        1. Try silent SSO using the Windows account broker (WAM).
           If your Windows session is already authenticated to your org AD,
           this succeeds with zero prompts.
        2. If the silent attempt fails (first run / token expired),
           a small browser popup appears once — you log in, token is cached
           locally by MSAL, and all future runs are silent again.

    Config required:
        sharepoint.client_id  — Azure AD App Registration client ID
                                (your IT team registers this once; it only needs
                                 Sites.ReadWrite.All *Delegated* permission)
        sharepoint.tenant_id  — your Azure AD tenant ID (or "common")
        sharepoint.site_url   — e.g. https://contoso.sharepoint.com/sites/MyTeam
        sharepoint.library    — Document Library name, e.g. "Documents"
        sharepoint.folder     — Subfolder path, e.g. "Jira Reports/SEC"
        sharepoint.filename   — exact filename to overwrite in SharePoint
    """

    GRAPH_BASE = "https://graph.microsoft.com/v1.0"
    SCOPES     = ["https://graph.microsoft.com/Sites.ReadWrite.All"]
    # MSAL caches the token here so silent auth works on subsequent runs
    TOKEN_CACHE_FILE = Path.home() / ".jira_extractor_token_cache.bin"

    def __init__(self, cfg: dict, logger: logging.Logger):
        sp               = cfg["sharepoint"]
        self.client_id   = sp["client_id"]
        self.tenant_id   = sp["tenant_id"]
        self.site_url    = sp["site_url"].rstrip("/")
        self.library     = sp["library"]
        self.folder      = sp["folder"].strip("/")
        self.sp_filename = sp.get("filename", cfg["project"]["output"])
        self.logger      = logger

    def _load_cache(self) -> msal.SerializableTokenCache:
        cache = msal.SerializableTokenCache()
        if self.TOKEN_CACHE_FILE.exists():
            cache.deserialize(self.TOKEN_CACHE_FILE.read_text(encoding="utf-8"))
        return cache

    def _save_cache(self, cache: msal.SerializableTokenCache) -> None:
        if cache.has_state_changed:
            self.TOKEN_CACHE_FILE.write_text(
                cache.serialize(), encoding="utf-8"
            )
            self.logger.debug("Token cache saved → %s", self.TOKEN_CACHE_FILE)

    def _get_token(self) -> str:
        """
        Acquire a Graph API token using the current Windows login session.
        Silent on every run after the first browser login.
        """
        cache     = self._load_cache()
        authority = f"https://login.microsoftonline.com/{self.tenant_id}"

        app = msal.PublicClientApplication(
            self.client_id,
            authority=authority,
            token_cache=cache,
        )

        # 1 — Try silent acquisition from cached token
        accounts = app.get_accounts()
        result   = None
        if accounts:
            self.logger.debug("Attempting silent token acquisition for: %s",
                              accounts[0].get("username"))
            result = app.acquire_token_silent(self.SCOPES, account=accounts[0])

        # 2 — Try Windows integrated auth (uses current Windows AD session, no popup)
        if not result or "access_token" not in result:
            self.logger.debug("Attempting Windows Integrated Auth (SSO) …")
            try:
                result = app.acquire_token_by_integrated_windows_auth(
                    scopes=self.SCOPES
                )
            except Exception:
                result = None

        # 3 — Fall back to interactive browser login (first run / token expired)
        if not result or "access_token" not in result:
            self.logger.info(
                "Opening browser for one-time SharePoint login "
                "(future runs will be silent) …"
            )
            result = app.acquire_token_interactive(scopes=self.SCOPES)

        if not result or "access_token" not in result:
            err = (result or {}).get("error_description",
                                     (result or {}).get("error", "Unknown"))
            raise RuntimeError(f"SharePoint authentication failed: {err}")

        self._save_cache(cache)
        self.logger.debug("SharePoint token acquired")
        return result["access_token"]

    def _get_site_id(self, token: str) -> str:
        from urllib.parse import urlparse
        parsed    = urlparse(self.site_url)
        hostname  = parsed.netloc
        site_path = parsed.path.strip("/")
        resp = requests.get(
            f"{self.GRAPH_BASE}/sites/{hostname}:/{site_path}",
            headers={"Authorization": f"Bearer {token}"},
            timeout=30,
        )
        resp.raise_for_status()
        site_id = resp.json()["id"]
        self.logger.debug("Resolved SharePoint site ID: %s", site_id)
        return site_id

    def _get_drive_id(self, token: str, site_id: str) -> str:
        resp = requests.get(
            f"{self.GRAPH_BASE}/sites/{site_id}/drives",
            headers={"Authorization": f"Bearer {token}"},
            timeout=30,
        )
        resp.raise_for_status()
        drives = resp.json().get("value", [])
        for drive in drives:
            if drive.get("name", "").lower() == self.library.lower():
                self.logger.debug("Resolved drive: %s → %s",
                                  self.library, drive["id"])
                return drive["id"]
        self.logger.warning("Library '%s' not found, using default drive",
                            self.library)
        return drives[0]["id"] if drives else "root"

    def upload(self, local_path: str) -> str:
        """
        Overwrite the existing file in the SharePoint folder.
        Creates the file if it does not exist yet.
        Returns the SharePoint web URL of the uploaded file.
        """
        file_path = Path(local_path)
        if not file_path.exists():
            raise FileNotFoundError(f"Local file not found: {file_path}")

        file_size = file_path.stat().st_size
        self.logger.info(
            "Uploading '%s' (%s KB) → %s / %s / %s",
            self.sp_filename, round(file_size / 1024, 1),
            self.site_url, self.library, self.folder,
        )

        token    = self._get_token()
        site_id  = self._get_site_id(token)
        drive_id = self._get_drive_id(token, site_id)

        dest      = f"{self.folder}/{self.sp_filename}" if self.folder else self.sp_filename
        auth_hdr  = {"Authorization": f"Bearer {token}"}

        # Always use upload session — works for any file size and guarantees overwrite
        session_url = (
            f"{self.GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
            f"/root:/{dest}:/createUploadSession"
        )
        sess = requests.post(
            session_url,
            headers={**auth_hdr, "Content-Type": "application/json"},
            json={"item": {"@microsoft.graph.conflictBehavior": "replace"}},
            timeout=30,
        )
        sess.raise_for_status()
        upload_url = sess.json()["uploadUrl"]

        CHUNK = 4 * 1024 * 1024   # 4 MB chunks
        uploaded = 0
        resp     = None

        with file_path.open("rb") as f:
            while True:
                chunk = f.read(CHUNK)
                if not chunk:
                    break
                end  = uploaded + len(chunk) - 1
                resp = requests.put(
                    upload_url,
                    headers={
                        "Content-Length": str(len(chunk)),
                        "Content-Range":  f"bytes {uploaded}-{end}/{file_size}",
                    },
                    data=chunk,
                    timeout=120,
                )
                resp.raise_for_status()
                uploaded += len(chunk)

        web_url = (resp.json() if resp else {}).get("webUrl", dest)
        self.logger.info("SharePoint upload complete → %s", web_url)
        return web_url


def gf(issue: dict, *keys, default=""):
    """Safely traverse nested issue fields."""
    val = issue.get("fields", {})
    for k in keys:
        if not isinstance(val, dict):
            return default
        val = val.get(k)
    return val if val is not None else default


def first_value(issue: dict, field_ids: list[str], default=""):
    for fid in field_ids:
        val = gf(issue, fid)
        if val not in (None, "", {}):
            return val
    return default


def fmt_date(raw) -> str:
    if not raw:
        return ""
    try:
        return datetime.fromisoformat(str(raw)[:10]).strftime("%d-%b-%Y")
    except ValueError:
        return str(raw)[:10]


def get_assignee(issue: dict) -> str:
    a = gf(issue, "assignee")
    return a.get("displayName", "") if isinstance(a, dict) else ""


def get_priority(issue: dict) -> str:
    p = gf(issue, "priority")
    return p.get("name", "") if isinstance(p, dict) else ""


def get_status(issue: dict) -> str:
    return gf(issue, "status", "name")


def get_labels(issue: dict) -> str:
    lb = gf(issue, "labels")
    return ", ".join(lb) if isinstance(lb, list) else ""


def get_story_points(issue: dict, field_ids: list[str]):
    val = first_value(issue, field_ids)
    return val if val != "" else ""


def get_start_date(issue: dict, field_ids: list[str]) -> str:
    return fmt_date(first_value(issue, field_ids))


def get_due_date(issue: dict, field_ids: list[str]) -> str:
    return fmt_date(first_value(issue, field_ids))


def resolve_custom_field(issue: dict, field_id: str, field_type: str) -> str:
    """
    Extract a custom field value and coerce it to a display string
    based on its declared type in config.

    Supported types:
        text   — plain string or nested .value / .name
        date   — ISO date string → DD-MMM-YYYY
        number — numeric value as-is
        user   — Jira user object → displayName
        array  — list of strings or objects → comma-separated
    """
    raw = gf(issue, field_id)

    if raw in (None, "", {}, []):
        return ""

    if field_type == "date":
        return fmt_date(raw)

    if field_type == "number":
        return str(raw)

    if field_type == "user":
        if isinstance(raw, dict):
            return raw.get("displayName", raw.get("name", ""))
        return str(raw)

    if field_type == "array":
        if isinstance(raw, list):
            parts = []
            for item in raw:
                if isinstance(item, dict):
                    parts.append(item.get("value") or item.get("name") or
                                 item.get("displayName") or str(item))
                else:
                    parts.append(str(item))
            return ", ".join(parts)
        return str(raw)

    # text (default) — handle nested option objects
    if isinstance(raw, dict):
        return raw.get("value") or raw.get("name") or raw.get("displayName") or str(raw)

    return str(raw)


# ──────────────────────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────────────────────

C_EPIC_BG  = "4A154B"
C_EPIC_FG  = "FFFFFF"
C_HDR_BG   = "172B4D"
C_HDR_FG   = "FFFFFF"
C_STORY_A  = "EFF6FF"
C_STORY_B  = "DBEAFE"
C_STORY_FG = "1E3A5F"
C_DONE     = "00875A"
C_INPROG   = "FF8B00"
C_TODO     = "6B778C"

STATUS_COLOURS = {
    "done":        C_DONE,
    "closed":      C_DONE,
    "resolved":    C_DONE,
    "in progress": C_INPROG,
    "in review":   C_INPROG,
}

# Fixed columns — custom_fields are appended after these at runtime
FIXED_COLS = [
    ("Epic",          22),   # A — merged per epic
    ("Story Key",     12),   # B
    ("Story Summary", 52),   # C
    ("Status",        14),   # D
    ("Assignee",      22),   # E
    ("Priority",      12),   # F
    ("Story Points",  13),   # G
    ("Start Date",    14),   # H
    ("Due Date",      14),   # I
    ("Labels",        26),   # J
]


def mk_fill(hex_col: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_col)


def mk_border_bottom() -> Border:
    s = Side(style="thin", color="CCCCCC")
    n = Side(style=None)
    return Border(bottom=s, top=n, left=n, right=n)


def mk_border_all() -> Border:
    s = Side(style="thin", color="BBBBBB")
    return Border(top=s, bottom=s, left=s, right=s)


def status_colour(status: str) -> str:
    return STATUS_COLOURS.get(status.lower(), C_TODO)


# ──────────────────────────────────────────────────────────────────────────────
# Workbook builder
# ──────────────────────────────────────────────────────────────────────────────

def build_workbook(epics: list, story_map: dict,
                   project_key: str, fields_cfg: dict,
                   custom_fields: list[dict]) -> Workbook:
    """
    custom_fields: list of dicts from config, e.g.
        [{"id": "customfield_10050", "label": "Team", "type": "text", "width": 18}, ...]
    """
    # Build full column list: fixed + custom
    cols = list(FIXED_COLS)
    for cf in custom_fields:
        width = int(cf.get("width", 20))
        cols.append((cf["label"], width))
    num_cols = len(cols)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{project_key} Hierarchy"
    ws.freeze_panes = "B3"

    for i, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 20

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws.cell(row=1, column=1,
                 value=f"Jira Project: {project_key}  —  Issue Hierarchy")
    tc.font      = Font(name="Arial", bold=True, size=12, color=C_HDR_FG)
    tc.fill      = mk_fill(C_HDR_BG)
    tc.alignment = Alignment(horizontal="center", vertical="center")

    # Header row
    for ci, (hdr, _) in enumerate(cols, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, size=10, color=C_HDR_FG)
        c.fill      = mk_fill(C_HDR_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = mk_border_bottom()

    sp_fields    = fields_cfg.get("story_points", ["customfield_10016"])
    start_fields = fields_cfg.get("start_date",   ["customfield_10015"])
    due_fields   = fields_cfg.get("due_date",      ["duedate"])

    current_row = 3

    for epic in epics:
        epic_key   = epic["key"]
        epic_name  = gf(epic, "summary")
        stories    = story_map.get(epic_key, [])
        epic_span  = max(1, len(stories))
        epic_start = current_row

        # Col A — Epic cell, merged across all story rows
        if epic_span > 1:
            ws.merge_cells(
                start_row=epic_start, start_column=1,
                end_row=epic_start + epic_span - 1, end_column=1,
            )
        ec = ws.cell(row=epic_start, column=1,
                     value=f"{epic_key}\n{epic_name}")
        ec.font      = Font(name="Arial", bold=True, size=10, color=C_EPIC_FG)
        ec.fill      = mk_fill(C_EPIC_BG)
        ec.alignment = Alignment(horizontal="center", vertical="center",
                                 wrap_text=True)
        ec.border    = mk_border_all()

        if not stories:
            for ci in range(2, num_cols + 1):
                c = ws.cell(row=epic_start, column=ci)
                c.fill   = mk_fill(C_EPIC_BG)
                c.border = mk_border_bottom()
            ws.row_dimensions[epic_start].height = 30
            current_row += 1
            continue

        for s_idx, story in enumerate(stories):
            row    = epic_start + s_idx
            bg     = C_STORY_A if s_idx % 2 == 0 else C_STORY_B
            status = get_status(story)
            ws.row_dimensions[row].height = 18

            def sc(col, val="", bold=False, center=False):
                c = ws.cell(row=row, column=col, value=val)
                c.font      = Font(name="Arial", size=10,
                                   color=C_STORY_FG, bold=bold)
                c.fill      = mk_fill(bg)
                c.alignment = Alignment(
                    horizontal="center" if center else "left",
                    vertical="center", wrap_text=True,
                )
                c.border = mk_border_bottom()
                return c

            # ── Fixed columns ─────────────────
            sc(2,  story["key"],                            bold=True)
            sc(3,  gf(story, "summary"))
            # Status badge
            stc = ws.cell(row=row, column=4, value=status)
            stc.font      = Font(name="Arial", bold=True, size=9, color="FFFFFF")
            stc.fill      = mk_fill(status_colour(status))
            stc.alignment = Alignment(horizontal="center", vertical="center")
            stc.border    = mk_border_bottom()

            sc(5,  get_assignee(story))
            sc(6,  get_priority(story),                     center=True)
            sc(7,  get_story_points(story, sp_fields),      center=True)
            sc(8,  get_start_date(story, start_fields),     center=True)
            sc(9,  get_due_date(story, due_fields),         center=True)
            sc(10, get_labels(story))

            # ── Dynamic custom field columns ──
            for cf_idx, cf in enumerate(custom_fields):
                col   = 11 + cf_idx          # starts at column K
                value = resolve_custom_field(story, cf["id"], cf.get("type", "text"))
                sc(col, value, center=(cf.get("type") in ("date", "number")))

        current_row = epic_start + epic_span

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 20

    summary_rows = [
        ("Metric",                "Value"),
        ("Total Epics",           len(epics)),
        ("Total Stories",         sum(len(v) for v in story_map.values())),
        ("Epics with no Stories", sum(1 for e in epics
                                      if not story_map.get(e["key"]))),
        ("Custom Fields",         ", ".join(cf["label"] for cf in custom_fields)
                                  if custom_fields else "None"),
        ("Extracted on",          datetime.now().strftime("%d-%b-%Y %H:%M")),
    ]
    for ri, (lbl, val) in enumerate(summary_rows, 1):
        lc = ws2.cell(row=ri, column=1, value=lbl)
        vc = ws2.cell(row=ri, column=2, value=val)
        if ri == 1:
            lc.font = vc.font = Font(name="Arial", bold=True,
                                     color=C_HDR_FG, size=10)
            lc.fill = vc.fill = mk_fill(C_HDR_BG)
        else:
            lc.font = vc.font = Font(name="Arial", size=10)
            vc.alignment = Alignment(horizontal="left", wrap_text=True)
        lc.alignment = Alignment(horizontal="left")
        ws2.row_dimensions[ri].height = 18

    return wb


# ──────────────────────────────────────────────────────────────────────────────
# Orchestrator
# ──────────────────────────────────────────────────────────────────────────────

def run(cfg: dict, logger: logging.Logger) -> None:
    client         = JiraClient(cfg, logger)
    project_key    = cfg["project"]["key"]
    output         = cfg["project"]["output"]
    fields_cfg     = cfg.get("fields", {})
    custom_fields  = cfg.get("custom_fields", [])

    if custom_fields:
        logger.info("Custom fields configured: %s",
                    ", ".join(f"{cf['label']} ({cf['id']})" for cf in custom_fields))

    # Build the full list of Jira field IDs to request
    base_fields = [
        "summary", "status", "assignee", "priority", "labels",
        *fields_cfg.get("story_points", ["customfield_10016", "customfield_10028"]),
        *fields_cfg.get("start_date",   ["customfield_10015", "startDate"]),
        *fields_cfg.get("due_date",     ["duedate", "customfield_10021"]),
        *[cf["id"] for cf in custom_fields],   # ← custom field IDs injected here
    ]
    base_fields = list(dict.fromkeys(base_fields))   # de-duplicate, preserve order
    logger.debug("Requesting fields: %s", base_fields)

    logger.info("Fetching Epics for project: %s", project_key)
    epics = client.fetch_all(
        f'project="{project_key}" AND issuetype=Epic ORDER BY key ASC',
        base_fields,
    )
    logger.info("Epics found: %d", len(epics))

    logger.info("Fetching Stories …")
    epic_link_fields = fields_cfg.get("epic_link", ["customfield_10014"])
    stories = client.fetch_all(
        f'project="{project_key}" AND issuetype=Story ORDER BY key ASC',
        base_fields + epic_link_fields + ["parent"],
    )
    logger.info("Stories found: %d", len(stories))

    # Map stories → epics
    story_map = {e["key"]: [] for e in epics}
    orphans   = []

    for story in stories:
        f = story.get("fields", {})
        epic_link = None
        for fid in epic_link_fields:
            epic_link = f.get(fid)
            if epic_link:
                break
        if not epic_link:
            epic_link = (f.get("parent") or {}).get("key")

        if epic_link and epic_link in story_map:
            story_map[epic_link].append(story)
        else:
            orphans.append(story)

    if orphans:
        logger.warning("%d stories have no linked Epic — grouped separately",
                       len(orphans))
        fake_epic = {
            "key": "NO-EPIC",
            "fields": {
                "summary":  "⚠ Stories without an Epic",
                "status":   {"name": "N/A"},
                "assignee": None, "priority": None, "labels": [],
            },
        }
        epics.append(fake_epic)
        story_map["NO-EPIC"] = orphans

    logger.info("Building workbook …")
    wb = build_workbook(epics, story_map, project_key, fields_cfg, custom_fields)
    wb.save(output)
    logger.info("Saved → %s  (Epics: %d | Stories: %d | Custom columns: %d)",
                output, len(epics),
                sum(len(v) for v in story_map.values()),
                len(custom_fields))

    # ── SharePoint upload ─────────────────────────────────────────────────────
    sp_cfg = cfg.get("sharepoint", {})
    if sp_cfg.get("enabled", False):
        sp_client = SharePointClient(cfg, logger)
        web_url   = sp_client.upload(output)
        logger.info("File available at: %s", web_url)
    else:
        logger.info("SharePoint upload skipped (sharepoint.enabled=false)")


# ──────────────────────────────────────────────────────────────────────────────
# Entry point
# ──────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract Jira Epics & Stories to a formatted Excel file."
    )
    parser.add_argument(
        "--config", "-c",
        default="config.json",
        help="Path to JSON config file (default: config.json)",
    )
    args = parser.parse_args()

    try:
        cfg    = load_config(args.config)
        logger = setup_logging(cfg)
        logger.info("Config loaded from: %s", args.config)
        run(cfg, logger)
    except FileNotFoundError as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except ValueError as e:
        print(f"CONFIG ERROR: {e}", file=sys.stderr)
        sys.exit(1)
    except requests.exceptions.ConnectionError:
        logging.error("Cannot connect to Jira. Check 'jira.url' in config.")
        sys.exit(1)
    except requests.exceptions.HTTPError as e:
        logging.error("Jira API error: %s", e.response.text)
        sys.exit(1)
    except Exception as e:
        logging.exception("Unexpected error: %s", e)
        sys.exit(1)


if __name__ == "__main__":
    main()
