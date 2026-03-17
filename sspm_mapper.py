"""
SSPM Policy Mapper — v4
========================================================
Key design decisions:
  - Pure semantic matching: NO framework tags injected into policy text
    (policies don't contain framework references — that's the point of this tool)
  - One app analysed at a time
  - Full relationship tracking: 1→many, many→1, many→many
  - Bidirectional gap analysis:
      * Controls with NO matching policy  → "Uncovered Controls"
      * Policies with NO matching control → "Orphan Policies"
  - Security domain flows through from input to output
"""

import json, csv, pickle, hashlib, warnings
import numpy as np
from pathlib import Path
from dataclasses import dataclass, field, asdict
from typing import Optional
warnings.filterwarnings("ignore")

CACHE_DIR = Path("policy_cache")
CACHE_DIR.mkdir(exist_ok=True)

# ── Thresholds ────────────────────────────────────────────────────────────────
THRESHOLD_FULL     = 0.82   # Strong match → FULL coverage
THRESHOLD_PARTIAL  = 0.60   # Moderate match → PARTIAL coverage
THRESHOLD_INDIRECT = 0.40   # Weak match → INDIRECT coverage
THRESHOLD_MIN      = 0.40   # Below this → no match at all

# ── Data Structures ───────────────────────────────────────────────────────────

@dataclass
class Control:
    """One row from the standards/controls input."""
    control_id:   str
    control_text: str
    domain:       str          # e.g. "Access Control", "Cryptography"
    framework:    str          # e.g. "ISO 27001", "NIST CSF"
    subdomain:    str = ""     # optional finer grain, e.g. "A.9.4"
    description:  str = ""     # optional longer description

@dataclass
class Policy:
    """One row from the OOTB policy input for a given app."""
    policy_id:    str
    policy_name:  str
    category:     str = ""     # vendor category if available (not framework)
    description:  str = ""     # vendor description if available
    impact:       str = ""     # impact level or description from policy file

@dataclass
class PolicyMatch:
    """A single (control → policy) match."""
    policy_id:        str
    policy_name:      str
    policy_category:  str
    similarity_score: float
    coverage:         str      # FULL / PARTIAL / INDIRECT
    impact:           str = "" # impact value carried from policy file
    confidence:       str = "" # HIGH / MEDIUM / LOW — match confidence flag

@dataclass
class ControlResult:
    """Mapping result for one control."""
    control_id:    str
    control_text:  str
    domain:        str
    framework:     str
    subdomain:     str
    risk_level:    str
    matches:       list = field(default_factory=list)   # list of PolicyMatch
    is_covered:    bool = False   # True if at least one FULL, PARTIAL or INDIRECT match

@dataclass
class PolicyResult:
    """Reverse-mapping result for one policy."""
    policy_id:       str
    policy_name:     str
    policy_category: str
    matched_controls: list = field(default_factory=list)  # list of control_ids
    is_orphan:       bool = True   # True if no control maps to this policy


# ── Loaders ───────────────────────────────────────────────────────────────────

class ControlLoader:
    """
    Load controls/standards from CSV or JSON.

    CSV columns (flexible naming):
      control_id, control_text, domain, framework, [subdomain], [description]

    JSON: list of objects with same fields, or
          { "domain": [...controls...] } grouped format
    """

    @staticmethod
    def from_csv(path: str) -> list[Control]:
        controls = []
        # utf-8-sig automatically strips BOM added by Excel when saving as CSV
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            # Detect ID column name once from headers
            id_col = None
            id_candidates = ["control_id","id","control_no","ctrl_id","ctr_id",
                             "no","number","ref","reference","control_ref",
                             "standard_id","req_id","requirement_id","ctrl_no"]
            fnames_lower = [f.strip().lower() for f in (reader.fieldnames or [])]
            for cand in id_candidates:
                if cand in fnames_lower:
                    id_col = cand
                    break
            if not id_col:
                print(f"ℹ️   Controls: no ID column found — will auto-generate IDs.")
                print(f"    Add one of these columns to use your own IDs:")
                print(f"    {id_candidates[:6]}")

            for i, row in enumerate(reader):
                row = {k.strip().lower(): v.strip() for k, v in row.items()}
                text = (row.get("control_text") or row.get("control") or
                        row.get("requirement") or row.get("description") or "")
                if not text or text.startswith("#"):
                    continue
                # Use detected ID column, else auto-generate
                if id_col:
                    ctrl_id = row.get(id_col) or f"CTR-{i+1:03d}"
                else:
                    ctrl_id = f"CTR-{i+1:03d}"
                controls.append(Control(
                    control_id   = ctrl_id,
                    control_text = text,
                    domain       = (row.get("domain") or row.get("security_domain") or
                                    row.get("category") or "General"),
                    framework    = (row.get("framework") or row.get("standard") or ""),
                    subdomain    = row.get("subdomain", row.get("sub_domain", "")),
                    description  = row.get("detail", row.get("notes", "")),
                ))
        return controls

    @staticmethod
    def from_json(path: str) -> list[Control]:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        controls = []
        if isinstance(raw, list):
            for i, item in enumerate(raw):
                # Try all accepted ID field names
                ctrl_id_val = (item.get("control_id") or item.get("id") or
                               item.get("control_no") or item.get("ctrl_id") or
                               item.get("no") or item.get("number") or
                               item.get("ref") or item.get("reference") or
                               f"CTR-{i+1:03d}")
                controls.append(Control(
                    control_id   = ctrl_id_val,
                    control_text = item.get("control_text", item.get("control", item.get("text", ""))),
                    domain       = item.get("domain", item.get("security_domain", "General")),
                    framework    = item.get("framework", item.get("standard", "")),
                    subdomain    = item.get("subdomain", ""),
                    description  = item.get("description", ""),
                ))
        elif isinstance(raw, dict):
            # { "Access Control": [{"control_id":..., "control_text":...}, ...] }
            for domain, items in raw.items():
                for i, item in enumerate(items):
                    controls.append(Control(
                        control_id   = (item.get("control_id") or item.get("id") or
                                        item.get("control_no") or item.get("ref") or
                                        f"{domain[:3].upper()}-{i+1:03d}"),
                        control_text = item.get("control_text", item.get("control", "")),
                        domain       = domain,
                        framework    = item.get("framework", ""),
                        subdomain    = item.get("subdomain", ""),
                        description  = item.get("description", ""),
                    ))
        return [c for c in controls if c.control_text]

    @staticmethod
    def load(path: str) -> list[Control]:
        ext = Path(path).suffix.lower()
        if ext == ".csv":  return ControlLoader.from_csv(path)
        if ext == ".json": return ControlLoader.from_json(path)
        raise ValueError(f"Unsupported format '{ext}'")


class PolicyLoader:
    """
    Load OOTB policies for ONE app from CSV or JSON.
    Policies must NOT contain framework references — pure vendor policy text only.
    """

    @staticmethod
    def from_csv(path: str, app_filter: Optional[str] = None) -> list[Policy]:
        policies = []
        # utf-8-sig automatically strips BOM added by Excel when saving as CSV
        with open(path, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            # Detect once from headers
            fieldnames_lower = [f.strip().lower() for f in (reader.fieldnames or [])]
            has_app_col = any(c in fieldnames_lower for c in ("app_name","app","application"))

            # Detect policy ID column name
            pol_id_col = None
            pol_id_candidates = ["policy_id","id","pol_id","policy_no","pol_no",
                                  "no","number","ref","policy_ref","ootb_id",
                                  "control_id","rule_id","check_id","finding_id"]
            for cand in pol_id_candidates:
                if cand in fieldnames_lower:
                    pol_id_col = cand
                    break
            if not pol_id_col:
                print(f"ℹ️   Policies: no ID column found — will auto-generate IDs.")
                print(f"    Add one of these columns to use your own IDs:")
                print(f"    {pol_id_candidates[:6]}")

            # Detect policy name column name
            pol_name_candidates = ["policy_name","policy","name","control",
                                   "rule","check","finding","title","description"]

            for i, row in enumerate(reader):
                row = {k.strip().lower(): v.strip() for k, v in row.items()}
                # Only filter by app name when the CSV actually has an app column
                if app_filter and has_app_col:
                    app_col = (row.get("app_name") or row.get("app") or
                               row.get("application") or "")
                    if app_col.lower() != app_filter.lower():
                        continue
                name = ""
                for nc in pol_name_candidates:
                    name = row.get(nc, "")
                    if name:
                        break
                if not name or name.startswith("#"):
                    continue
                # Use detected ID column, else auto-generate
                if pol_id_col:
                    pol_id = row.get(pol_id_col) or f"POL-{i+1:03d}"
                else:
                    pol_id = f"POL-{i+1:03d}"
                policies.append(Policy(
                    policy_id   = pol_id,
                    policy_name = name,
                    category    = row.get("category", row.get("type", "")),
                    description = row.get("description", row.get("desc", "")),
                    impact      = row.get("impact", row.get("impact_level",
                                  row.get("severity", row.get("priority", "")))),
                ))
        return policies

    @staticmethod
    def from_json(path: str, app_filter: Optional[str] = None) -> list[Policy]:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        policies = []
        items = []
        if isinstance(raw, list):
            items = raw
        elif isinstance(raw, dict):
            if app_filter and app_filter in raw:
                items = raw[app_filter]
            elif app_filter:
                # Try case-insensitive match
                for k, v in raw.items():
                    if k.lower() == app_filter.lower():
                        items = v
                        break
            else:
                # Flatten all
                for v in raw.values():
                    items.extend(v if isinstance(v, list) else [v])
        for i, item in enumerate(items):
            if isinstance(item, str):
                policies.append(Policy(
                    policy_id=f"POL-{i+1:03d}", policy_name=item,
                    category="", description=""))
            else:
                pol_id_val = (item.get("policy_id") or item.get("id") or
                              item.get("pol_id") or item.get("policy_no") or
                              item.get("no") or item.get("number") or
                              item.get("ref") or item.get("rule_id") or
                              f"POL-{i+1:03d}")
                policies.append(Policy(
                    policy_id   = pol_id_val,
                    policy_name = item.get("policy_name", item.get("policy", item.get("name", ""))),
                    category    = item.get("category", ""),
                    description = item.get("description", ""),
                    impact      = item.get("impact", item.get("impact_level",
                                  item.get("severity", item.get("priority", "")))),
                ))
        return [p for p in policies if p.policy_name]

    @staticmethod
    def load(path: str, app_filter: Optional[str] = None) -> list[Policy]:
        ext = Path(path).suffix.lower()
        if ext == ".csv":  return PolicyLoader.from_csv(path, app_filter)
        if ext == ".json": return PolicyLoader.from_json(path, app_filter)
        raise ValueError(f"Unsupported format '{ext}'")


# ── Encoder ───────────────────────────────────────────────────────────────────

# ── SET THIS to your local model folder — NEVER leave as None in production ──
# Run python download_secbert.py once to download the model locally.
# Then set this path so the program NEVER contacts HuggingFace again.
#
# Examples:
#   SECBERT_MODEL_PATH = "./secbert_clean"        # relative to project folder
#   SECBERT_MODEL_PATH = "./secbert_model"        # if you used download_secbert.py
#   SECBERT_MODEL_PATH = "/opt/models/secbert"    # absolute path
#
# If None → attempts HuggingFace download (requires internet)
SECBERT_MODEL_PATH = "./secbert_clean"   # ← SET THIS TO YOUR FOLDER


def _find_model_path(base_path: str) -> str:
    """
    Resolve the actual model directory from a given base path.

    Handles two folder structures:
      1. Flat (save_pretrained output):
             secbert_clean/
               config.json
               model.safetensors / pytorch_model.bin
               tokenizer.json
               vocab.txt

      2. HuggingFace cache structure (cache_dir output):
             secbert_model/
               models--jackaduma--SecBERT/
                 snapshots/
                   <hash>/
                     config.json ...

    Returns the path that actually contains config.json.
    Raises FileNotFoundError with a clear message if nothing found.
    """
    import os
    from pathlib import Path

    base = Path(base_path)

    # ── Check if base itself is a flat model folder ───────────────────────────
    if (base / "config.json").exists():
        return str(base)

    # ── Check HuggingFace cache structure ─────────────────────────────────────
    # Pattern: base/models--<org>--<model>/snapshots/<hash>/config.json
    for snapshots_dir in base.rglob("snapshots"):
        if snapshots_dir.is_dir():
            # Get the most recent snapshot (sorted by name)
            snapshots = sorted(snapshots_dir.iterdir(), reverse=True)
            for snap in snapshots:
                if (snap / "config.json").exists():
                    return str(snap)

    # ── Check one level deep (user may have pointed to parent folder) ─────────
    for child in base.iterdir():
        if child.is_dir() and (child / "config.json").exists():
            return str(child)

    # ── Nothing found ─────────────────────────────────────────────────────────
    raise FileNotFoundError(
        f"\n\n  ❌  SecBERT model not found at: {base_path}"
        f"\n     Expected to find config.json inside this folder."
        f"\n     Run:  python download_secbert.py"
        f"\n     Then set SECBERT_MODEL_PATH = \"./secbert_clean\" in sspm_mapper.py"
    )


class SecBERTEncoder:
    """
    Encodes text using SecBERT.

    Loading priority:
      1. model_path argument (passed at runtime via --model)
      2. SECBERT_MODEL_PATH constant set at top of this file
      3. HuggingFace download — ONLY if both above are None/empty

    Automatically handles both flat (save_pretrained) and
    HuggingFace cache_dir folder structures.
    Falls back to TF-IDF if transformers/torch unavailable.
    """

    def __init__(self, model_path: str = None):
        self.mode = None
        # Priority: runtime arg > file constant > HuggingFace (last resort)
        raw_path = model_path or SECBERT_MODEL_PATH or None
        self.model_path = raw_path   # may be None if no local path configured
        self._load()

    def _load(self):
        try:
            from transformers import AutoTokenizer, AutoModel
            import torch

            if self.model_path:
                # ── Local folder — resolve exact path, never hit network ──────
                resolved = _find_model_path(self.model_path)
                print(f"🔄  Loading SecBERT from local folder: {resolved}")
                self.tokenizer = AutoTokenizer.from_pretrained(
                    resolved, local_files_only=True)
                self.model = AutoModel.from_pretrained(
                    resolved, local_files_only=True)
            else:
                # ── No local path set — download from HuggingFace ─────────────
                print(f"⚠️   SECBERT_MODEL_PATH is not set.")
                print(f"     Downloading from HuggingFace (requires internet)...")
                print(f"     To avoid this, run: python download_secbert.py")
                print(f"     Then set SECBERT_MODEL_PATH = \"./secbert_clean\" in sspm_mapper.py\n")
                self.tokenizer = AutoTokenizer.from_pretrained("jackaduma/SecBERT")
                self.model     = AutoModel.from_pretrained("jackaduma/SecBERT")

            self.model.eval()
            self.torch = torch
            self.mode  = "secbert"

            # Use GPU if available — significant speedup for large batches
            if torch.cuda.is_available():
                self._device = torch.device("cuda")
                self.model   = self.model.to(self._device)
                print(f"⚡  GPU detected — using CUDA for encoding")
            elif torch.backends.mps.is_available():
                # Apple Silicon GPU
                self._device = torch.device("mps")
                self.model   = self.model.to(self._device)
                print(f"⚡  Apple Silicon GPU (MPS) detected")
            else:
                self._device = torch.device("cpu")
                print(f"ℹ️   Running on CPU — encoding will take ~30-60s for first run")

            print("✅  SecBERT ready.\n")

        except FileNotFoundError as e:
            print(e)
            print("\n    → Falling back to TF-IDF mode\n")
            self.mode = "tfidf"
        except Exception as e:
            print(f"⚠️  SecBERT unavailable ({e})\n    → TF-IDF fallback mode\n")
            self.mode = "tfidf"

    def fit_tfidf(self, corpus: list[str]):
        from sklearn.feature_extraction.text import TfidfVectorizer
        self.vectorizer = TfidfVectorizer(ngram_range=(1, 3), stop_words="english")
        self.vectorizer.fit(corpus)

    def encode(self, texts: list[str]) -> np.ndarray:
        if self.mode == "secbert":
            return self._encode_secbert(texts)
        return self.vectorizer.transform(texts).toarray()

    def _encode_secbert(self, texts: list[str],
                        batch_size: int = 32) -> np.ndarray:
        """
        Encode texts in batches — far faster than one-at-a-time.
        batch_size=32 works well on CPU. Increase to 64+ if you have GPU RAM.
        """
        import torch
        all_vecs = []

        with torch.no_grad():
            for start in range(0, len(texts), batch_size):
                batch = texts[start: start + batch_size]

                # Tokenize entire batch at once — padding to longest in batch
                inputs = self.tokenizer(
                    batch,
                    return_tensors="pt",
                    truncation=True,
                    max_length=256,
                    padding=True,       # pad shorter sequences in batch
                )

                # Move to GPU if available
                if hasattr(self, '_device'):
                    inputs = {k: v.to(self._device) for k, v in inputs.items()}

                outputs = self.model(**inputs)
                hidden  = outputs.last_hidden_state          # (B, seq, hidden)
                mask    = inputs["attention_mask"].unsqueeze(-1).float()
                pooled  = (hidden * mask).sum(1) / mask.sum(1)  # (B, hidden)

                # Move back to CPU for numpy
                vecs = pooled.cpu().numpy()                  # (B, hidden)
                all_vecs.append(vecs)

        return np.vstack(all_vecs)   # (N, hidden)


# ── Cache ─────────────────────────────────────────────────────────────────────

class EmbeddingCache:
    def _key(self, label: str, texts: list[str]) -> str:
        # Hash includes label + all policy text content
        # This means cache auto-invalidates when policy content changes
        return hashlib.sha256((label + "||".join(texts)).encode()).hexdigest()[:16]

    def get(self, label: str, texts: list[str]) -> Optional[np.ndarray]:
        p = CACHE_DIR / f"{self._key(label, texts)}.pkl"
        return pickle.load(open(p, "rb")) if p.exists() else None

    def set(self, label: str, texts: list[str], vecs: np.ndarray):
        pickle.dump(vecs, open(CACHE_DIR / f"{self._key(label, texts)}.pkl", "wb"))

    def clear(self):
        """Delete all cached embeddings. Run when switching model or policy files."""
        deleted = 0
        for p in CACHE_DIR.glob("*.pkl"):
            p.unlink()
            deleted += 1
        print(f"🗑️   Cache cleared ({deleted} files deleted).")


# ── Policy Text Builder ───────────────────────────────────────────────────────

# ── Security keyword groups for keyword-boost scoring ────────────────────────
# Tighter, non-overlapping groups. Each keyword only belongs to one group.
# Overlap between groups causes false boosts — keep groups distinct.
# ── Synonym expansion map ────────────────────────────────────────────────────
# Expands abbreviations/synonyms BEFORE encoding so TF-IDF treats them the same
# e.g. "MFA" → "MFA multi-factor authentication two-factor"
# Applied to BOTH control and policy text before vectorisation
SYNONYM_MAP = {
    # Authentication
    r"\bmfa\b":                   "mfa multi-factor authentication two-factor 2fa",
    r"\b2fa\b":                   "2fa two-factor authentication mfa multi-factor",
    r"\bsso\b":                   "sso single sign-on saml federation",
    r"\bsaml\b":                  "saml sso single sign-on federation",
    r"\boidc\b":                  "oidc openid connect oauth federation",
    r"\bfido\b":                  "fido fido2 hardware key security key phishing-resistant",
    r"\bpam\b":                   "pam privileged access management",
    r"\bpim\b":                   "pim privileged identity management just-in-time",
    # Access Control
    r"\brbac\b":                  "rbac role-based access control permission",
    r"\babac\b":                  "abac attribute-based access control",
    r"\bjit\b":                   "jit just-in-time access",
    r"\biam\b":                   "iam identity access management",
    # Data Protection
    r"\bdlp\b":                   "dlp data loss prevention sensitive data",
    r"\bpii\b":                   "pii personally identifiable information sensitive data",
    r"\bphi\b":                   "phi personal health information sensitive data",
    r"\bpci\b":                   "pci payment card industry financial data",
    r"\bgdpr\b":                  "gdpr data protection regulation privacy",
    r"\bhipaa\b":                 "hipaa health information privacy compliance",
    # Encryption
    r"\btls\b":                   "tls transport layer security ssl encrypt",
    r"\bssl\b":                   "ssl secure sockets layer tls encrypt",
    r"\baes\b":                   "aes advanced encryption standard encrypt",
    r"\bkms\b":                   "kms key management service encryption",
    r"\be2e\b":                   "e2e end-to-end encryption",
    # Infrastructure
    r"\bsiem\b":                  "siem security information event management log monitoring",
    r"\bsoc\b":                   "soc security operations center monitoring",
    r"\bueba\b":                  "ueba user entity behaviour analytics anomaly detection",
    r"\bcasb\b":                  "casb cloud access security broker",
    r"\bcspm\b":                  "cspm cloud security posture management",
    r"\bsspm\b":                  "sspm saas security posture management",
    r"\bmdm\b":                   "mdm mobile device management endpoint",
    r"\bbyod\b":                  "byod bring your own device mobile endpoint",
    r"\bvpn\b":                   "vpn virtual private network remote access",
    r"\bscim\b":                  "scim system cross-domain identity management provisioning",
    # Protocols blocked
    r"\bsmtp\b":                  "smtp email legacy authentication protocol",
    r"\bimap\b":                  "imap email legacy authentication protocol",
    r"\bpop3\b":                  "pop3 email legacy authentication protocol",
    r"\bntlm\b":                  "ntlm legacy windows authentication protocol",
    # Frameworks
    r"\bnist\b":                  "nist cybersecurity framework standard",
    r"\biso\b":                   "iso international standard",
    r"\bsox\b":                   "sox sarbanes oxley compliance financial",
    # Actions
    r"\bdecommission\b":         "decommission disable remove block",
    r"\bdeprovision\b":          "deprovision remove revoke disable account",
    r"\blegacy auth\b":            "legacy authentication basic auth old protocol",
    r"at rest":                    "at rest storage encryption data stored",
    r"in transit":                 "in transit network encryption data transmitted",
    r"step-up":                    "step-up authentication mfa challenge",
    r"least privilege":            "least privilege minimum access rbac permission",
    r"zero trust":                 "zero trust verify always access control",
    r"data residency":             "data residency geographic location region compliance",
    r"audit trail":                "audit trail log record activity history",
    r"anomalous":                  "anomalous unusual suspicious behaviour detection",
    # Explicit at-rest vs in-transit expansions
    r"at.rest":            "at rest storage encryption data stored encrypted storage",
    r"in.transit":         "in transit transport encryption tls ssl network encrypted",
    r"data.at.rest":       "data at rest storage encryption encrypted storage",
    r"data.in.transit":    "data in transit transport tls ssl network encryption",
    r"transparent.data":   "transparent data encryption at rest database storage",
}


# ── Concept Bridge Map ────────────────────────────────────────────────────────
# Maps HIGH-LEVEL control concepts → POLICY-LEVEL implementation vocabulary
# This is the key fix for mismatches like:
#   "geofencing" → "conditional access location-based policy"
#   "risky sign-in" → "identity protection risk-based conditional access"
#
# Each entry: (phrase_in_control, text_to_append_to_encoding)
# The appended text uses the vocabulary the POLICY would use
#
# Add your own entries here when you find bad/missing matches

CONCEPT_BRIDGES = [
    # ── Location & Network Access ─────────────────────────────────────────────
    ("geofenc",
     "geofencing location-based access conditional access trusted location "
     "ip range network location restriction country block"),

    ("location-based",
     "location-based access conditional access trusted location ip restriction "
     "named location country-based geofencing"),

    ("country.*block",
     "country block conditional access location policy geofencing ip restriction"),

    ("trusted location",
     "trusted location conditional access named location ip range geofencing"),

    ("network restriction",
     "network restriction ip range location conditional access allowlist"),

    # ── Risk-Based Access ─────────────────────────────────────────────────────
    ("risky sign",
     "risky sign-in risk-based conditional access identity protection "
     "risk policy anomalous login suspicious authentication block"),

    ("risky user",
     "risky user identity protection risk-based access conditional access "
     "compromised account high risk block"),

    ("risk-based access",
     "risk-based access conditional access identity protection sign-in risk "
     "user risk policy"),

    ("high risk",
     "high risk user sign-in risk identity protection conditional access "
     "block remediate"),

    ("compromised account",
     "compromised account risky user identity protection conditional access "
     "block revoke session"),

    ("anomalous login",
     "anomalous login risky sign-in identity protection conditional access "
     "unusual activity suspicious"),

    # ── Identity & Authentication ─────────────────────────────────────────────
    ("step-up auth",
     "step-up authentication mfa challenge conditional access sensitive resource"),

    ("adaptive auth",
     "adaptive authentication risk-based mfa conditional access"),

    ("passwordless",
     "passwordless fido2 security key windows hello authentication"),

    ("continuous auth",
     "continuous authentication session re-evaluation conditional access"),

    # ── Data Governance ───────────────────────────────────────────────────────
    ("data classification",
     "data classification sensitivity label information protection dlp policy"),

    ("sensitivity label",
     "sensitivity label data classification information protection dlp encrypt"),

    ("information barrier",
     "information barrier dlp communication compliance data segmentation"),

    ("data sovereignty",
     "data sovereignty data residency region compliance geographic restriction"),

    # ── Endpoint & Device ─────────────────────────────────────────────────────
    ("unmanaged device",
     "unmanaged device byod conditional access device compliance mdm"),

    ("device compliance",
     "device compliance conditional access managed device intune mdm"),

    ("jailbreak",
     "jailbreak rooted device mdm endpoint compliance block"),

    # ── Threat & Incident ─────────────────────────────────────────────────────
    ("lateral movement",
     "lateral movement privileged access pam session monitoring detection"),

    ("impossible travel",
     "impossible travel risky sign-in identity protection anomaly alert"),

    ("brute force",
     "brute force password lockout account protection smart lockout"),

    ("credential stuffing",
     "credential stuffing password spray smart lockout identity protection"),

    # ── Sharing & Collaboration ───────────────────────────────────────────────
    ("oversharing",
     "oversharing external sharing dlp data classification sensitivity label"),

    ("guest sharing",
     "guest sharing external user access review conditional access"),

    ("email forwarding",
     "email forwarding mailbox rule exfiltration dlp transport rule"),

    # ── Privilege & Admin ─────────────────────────────────────────────────────
    ("standing access",
     "standing access just-in-time pim privileged access time-bound"),

    ("emergency access",
     "emergency access break glass privileged account admin"),

    ("shadow admin",
     "shadow admin privileged account access review role assignment"),

    # ── Compliance ────────────────────────────────────────────────────────────
    ("right to be forgotten",
     "right to be forgotten gdpr data deletion retention policy"),

    ("data minimisation",
     "data minimisation gdpr privacy dlp classification retention"),

    ("cross-border transfer",
     "cross-border transfer data residency gdpr compliance geographic"),
]


# ── Vendor Policy Name Library ───────────────────────────────────────────────
# When a policy has NO description (common with OOTB policies), this library
# provides semantic descriptions automatically keyed on the policy name.
# This solves the "short policy name vs long control text" mismatch.
# Add entries for your specific vendor's policy names here.

POLICY_NAME_LIBRARY = {
    # ── Salesforce — Identity & Authentication ────────────────────────────────
    "enable sso": (
        "Enable Single Sign-On SSO SAML federation centralized authentication "
        "identity provider all applications must federate"
    ),
    "enforce authentication through custom domain": (
        "Enforce SSO authentication through custom domain SAML identity provider "
        "federated login prevent direct salesforce login"
    ),
    "disable login with salesforce credentials": (
        "Disable native Salesforce credential login enforce SSO single sign-on "
        "identity provider federated authentication"
    ),
    "mfa": (
        "Multi-factor authentication MFA two-factor step-up enforced all users "
        "login verification second factor"
    ),
    "physical security key authentication": (
        "Physical security key FIDO2 hardware token phishing-resistant MFA "
        "step-up authentication privileged accounts"
    ),
    "enable authenticator passwordless login": (
        "Authenticator passwordless login lightning login FIDO MFA two-factor "
        "authentication step-up"
    ),
    "certificate-based authentication": (
        "Certificate-based authentication PKI x509 mutual TLS strong authentication "
        "MFA alternative"
    ),
    "identity verification with sms": (
        "SMS identity verification two-factor MFA step-up authentication "
        "second factor verification code"
    ),
    "force non-email verification methods": (
        "Force non-email MFA verification methods authenticator app security key "
        "phishing-resistant step-up"
    ),
    "require identity verification for email address change": (
        "Require identity verification MFA step-up for email address change "
        "account modification sensitive change"
    ),
    "email confirmations for email change": (
        "Email confirmation for email address change identity verification "
        "account security external users communities"
    ),
    # ── Salesforce — Session Management ──────────────────────────────────────
    "inactive sessions logout timeout": (
        "Inactive session logout timeout idle session expiry auto-logout "
        "session management controls"
    ),
    "inactive sessions logout": (
        "Inactive session logout timeout idle expiry session management "
        "auto-logout session controls"
    ),
    "profile inactive sessions logout timeout": (
        "Profile-level inactive session logout timeout idle expiry "
        "session management per profile"
    ),
    "session timeouts": (
        "Session timeout idle expiry management controls auto-logout "
        "re-authentication session duration"
    ),
    "enforce session termination on admin password reset": (
        "Enforce session termination terminate all sessions on admin password reset "
        "security incident response revoke session"
    ),
    "lock sessions to the domain in which they were first used": (
        "Lock session to originating domain session security prevent hijacking "
        "session binding"
    ),
    "re-login after login as user": (
        "Re-login after admin login-as-user session security admin access "
        "audit logging"
    ),
    "external client application- required session level": (
        "External client application required session level assurance "
        "MFA session security OAuth"
    ),
    # ── Salesforce — Password & Lockout ──────────────────────────────────────
    "lockout interval": (
        "Lockout interval account lockout duration brute force protection "
        "failed login attempts password policy"
    ),
    "profile lockout interval": (
        "Profile lockout interval per-profile account lockout duration "
        "brute force protection"
    ),
    "invalid login attempts": (
        "Invalid login attempts threshold account lockout brute force protection "
        "failed authentication password policy"
    ),
    "security check": (
        "Security health check password policy settings baseline configuration "
        "security posture assessment"
    ),
    "autocomplete user name": (
        "Autocomplete username disable credential exposure browser autofill "
        "login security"
    ),
    "password reset mechanism": (
        "Password reset mechanism self-service password reset alignment "
        "identity management standards procedures"
    ),
    # ── Salesforce — Access Control & IP Restriction ──────────────────────────
    "trusted ip ranges configuration": (
        "Trusted IP ranges configuration geofencing location-based access "
        "conditional access network restriction allowlist whitelist "
        "restrict login by location country"
    ),
    "high privilege profile ip restriction": (
        "High privilege profile IP restriction location-based access "
        "geofencing privileged access admin restriction trusted network"
    ),
    "non privilege profile ip restriction": (
        "Non-privilege profile IP restriction location-based access "
        "geofencing network restriction trusted IP"
    ),
    "connected application ip restriction": (
        "Connected application IP restriction geofencing location-based "
        "network restriction OAuth app access control"
    ),
    "external client application- ip restriction enforcement": (
        "External client application IP restriction enforcement geofencing "
        "location-based access network restriction"
    ),
    "login from known ip ranges on every page request": (
        "Login from known IP ranges every page request continuous location "
        "verification geofencing conditional access network"
    ),
    "limit connected apps api access": (
        "Limit connected apps API access allowlisted third-party integration "
        "OAuth restriction API governance"
    ),
    "restrict customers and partners api access": (
        "Restrict customers partners API access third-party integration "
        "external API restriction access control"
    ),
    "limit connected apps to specific profiles or permission sets": (
        "Limit connected apps specific profiles permission sets RBAC "
        "role-based access OAuth app restriction"
    ),
    "oauth connected apps with full access scope": (
        "OAuth connected apps full access scope third-party integration "
        "admin consent API permission restriction"
    ),
    "oauth username-password flows": (
        "OAuth username password flows legacy authentication disable "
        "insecure OAuth flow restriction"
    ),
    "oauth user-agent flow": (
        "OAuth user agent flow implicit flow legacy insecure "
        "restriction OAuth security"
    ),
    "oauth pkce requirement": (
        "OAuth PKCE requirement proof key code exchange secure OAuth "
        "authorization code flow security"
    ),
    "oauth credential flow": (
        "OAuth credential flow client credentials machine-to-machine "
        "API authentication restriction"
    ),
    "connected app uses non expiring refresh tokens": (
        "Connected app non-expiring refresh tokens token expiry "
        "session management OAuth token lifecycle"
    ),
    "external client application- single logout": (
        "External client application single logout SLO SSO federation "
        "session termination"
    ),
    "external client application- secret requirement": (
        "External client application secret requirement OAuth client secret "
        "API authentication credential"
    ),
    "external client application - introspect all tokens": (
        "External client application introspect tokens OAuth token validation "
        "security"
    ),
    "external client application - client credential flow": (
        "External client application client credential flow OAuth machine "
        "API authentication restriction"
    ),
    "named credentials anonymous authentication": (
        "Named credentials anonymous authentication restrict third-party "
        "integration credential management"
    ),
    # ── Salesforce — User Provisioning & Lifecycle ────────────────────────────
    "dormant users": (
        "Dormant users inactive account deprovisioning user lifecycle "
        "account management disable stale accounts"
    ),
    "users with login access to support": (
        "Users with login access to support admin access review "
        "privileged access governance"
    ),
    "admins log-in to other user's account": (
        "Admins login to other user account privileged access audit "
        "admin capability governance logging"
    ),
    "self registration for digital workspace site": (
        "Self-registration digital workspace external user provisioning "
        "user lifecycle community sites"
    ),
    "prevent using standard external profiles for self-registration and user creation": (
        "Prevent standard external profiles self-registration user creation "
        "user lifecycle provisioning governance"
    ),
    "user account provisioning": (
        "User account provisioning deprovisioning lifecycle management "
        "identity standards SCIM automation"
    ),
    # ── Salesforce — Permissions & Role-Based Access ──────────────────────────
    "field-level security for flexcards using soql data sources enforcement": (
        "Field-level security FLS SOQL data sources enforcement least privilege "
        "data access control attribute-based ABAC"
    ),
    "access controls on cached integration procedure metadata enforcement": (
        "Access controls cached integration procedure metadata enforcement "
        "least privilege RBAC permission"
    ),
    "remote action execution to authorized apex classes restriction": (
        "Remote action execution authorized Apex classes restriction "
        "least privilege code execution access control"
    ),
    "require permission to view record names in lookup fields": (
        "Require permission view record names lookup fields least privilege "
        "field-level access control"
    ),
    "role- and attribute-base access controls": (
        "Role attribute-based access controls RBAC ABAC least privilege "
        "identity management standards"
    ),
    # ── Salesforce — Data & Metadata Security ────────────────────────────────
    "restrict access to custom metadata types": (
        "Restrict access custom metadata types data governance access control "
        "least privilege"
    ),
    "restrict access to custom settings": (
        "Restrict access custom settings configuration security least privilege "
        "admin restriction"
    ),
    "disable sosl search on custom settings": (
        "Disable SOSL search custom settings data exposure restriction "
        "access control"
    ),
    "field-level security and encryption controls in data mappers enforcement": (
        "Field-level security encryption controls data mappers enforcement "
        "data protection least privilege OmniStudio"
    ),
    "excessive access to omnistudio data pack attachments": (
        "Excessive access OmniStudio data pack attachments least privilege "
        "data access governance"
    ),
    "permissions for nested integration procedure actions enforcement": (
        "Permissions nested integration procedure actions enforcement "
        "least privilege access control OmniStudio"
    ),
    "disable scale cache to prevent unauthorized execution of data mappers": (
        "Disable scale cache prevent unauthorized execution data mappers "
        "security control access restriction OmniStudio"
    ),
    # ── Salesforce — Network & Protocol Security ─────────────────────────────
    "remote sites protocol (http/s) security": (
        "Remote sites protocol HTTP HTTPS security TLS in-transit encryption "
        "network security remote endpoint"
    ),
    "remote endpoints (remote sites) without tls": (
        "Remote endpoints without TLS insecure HTTP protocol in-transit "
        "encryption network security"
    ),
    "warn on redirection out of salesforce": (
        "Warn redirection out of Salesforce external redirect security "
        "phishing protection"
    ),
    "require httponly attribute": (
        "Require HttpOnly attribute cookie security session security "
        "XSS protection"
    ),
    # ── Salesforce — Community & External Access ──────────────────────────────
    "community sites - guest users public chatter api access": (
        "Community sites guest users public Chatter API access external "
        "user restriction least privilege Digital Experience"
    ),
    "community sites - guest files access": (
        "Community sites guest files access external user restriction "
        "least privilege Digital Experience"
    ),
    "public report folders accessible by all users": (
        "Public report folders accessible all users least privilege "
        "data exposure restriction access control"
    ),
    "public dashboard folders accessible by all users": (
        "Public dashboard folders accessible all users least privilege "
        "data exposure restriction"
    ),
    "customer invitations to private groups": (
        "Customer invitations private groups external access governance "
        "community access control"
    ),
    # ── Salesforce — Monitoring & Audit ──────────────────────────────────────
    "[shield] analytics - number of users with admin access to event monitoring": (
        "Shield analytics admin access event monitoring audit log "
        "privileged access logging"
    ),
    "[shield] analytics - number of users with read access to event monitoring": (
        "Shield analytics read access event monitoring audit logging "
        "security monitoring"
    ),
    "connected apps with user interface access": (
        "Connected apps user interface access OAuth third-party "
        "application access governance"
    ),
}


def enrich_policy_from_library(policy_name: str, existing_description: str) -> str:
    """
    Look up the policy name in POLICY_NAME_LIBRARY.
    If found AND the policy has no description (or a very short one),
    return the library description to augment encoding.
    Always appends library text — it never replaces a real description.
    """
    name_lower = policy_name.strip().lower()
    # Try exact match first
    if name_lower in POLICY_NAME_LIBRARY:
        lib_text = POLICY_NAME_LIBRARY[name_lower]
        if existing_description and len(existing_description.strip()) > 20:
            return existing_description + " " + lib_text
        return lib_text
    # Try partial match — if policy name contains a known key phrase
    for key, lib_text in POLICY_NAME_LIBRARY.items():
        if key in name_lower or name_lower in key:
            if existing_description and len(existing_description.strip()) > 20:
                return existing_description + " " + lib_text
            return lib_text
    return existing_description


def expand_concepts(text: str) -> str:
    """
    Expand high-level control concepts into policy-implementation vocabulary.
    Called AFTER expand_synonyms — adds bridging text at the end.
    Example:
      "geofencing must restrict access" →
      "geofencing must restrict access [geofencing location-based access
       conditional access trusted location ip range ...]"
    """
    import re
    t     = text.lower()
    extra = []
    for phrase, expansion in CONCEPT_BRIDGES:
        if re.search(phrase, t, flags=re.IGNORECASE):
            extra.append(expansion)
    if extra:
        return text + " " + " ".join(extra)
    return text


def expand_synonyms(text: str) -> str:
    """
    Expand abbreviations and synonyms to improve TF-IDF matching.
    Applied once — expansions are not re-expanded to avoid chain substitution.
    """
    import re
    result = text.lower()
    # Collect replacements first, apply all at once to avoid chain-expanding
    expanded_positions = set()
    replacements = []
    for pattern, expansion in SYNONYM_MAP.items():
        for m in re.finditer(pattern, result, flags=re.IGNORECASE):
            # Skip if this position was already expanded by a previous pattern
            if any(m.start() < ep[1] and m.end() > ep[0]
                   for ep in expanded_positions):
                continue
            replacements.append((m.start(), m.end(), expansion))
            expanded_positions.add((m.start(), m.end()))

    # Apply replacements from end to start to preserve positions
    replacements.sort(key=lambda x: x[0], reverse=True)
    for start, end, expansion in replacements:
        result = result[:start] + expansion + result[end:]
    return result


KEYWORD_GROUPS = {
    # ── Authentication & Identity ─────────────────────────────────────────────
    "mfa": [
        "mfa","multi-factor","two-factor","2fa","step-up","step up auth",
        "hardware token","fido","fido2","authenticator app","otp",
        "one-time password","totp","phishing-resistant","passwordless",
        "security key","verification code","second factor","strong auth",
        "adaptive auth","risk-based auth","challenge","verify identity"
    ],
    "legacy_auth": [
        "legacy auth","basic auth","smtp auth","imap","pop3","ntlm","kerberos",
        "disable authentication","block protocol","old protocol","legacy protocol",
        "deprecated auth","insecure protocol","cleartext","unencrypted auth"
    ],
    "sso_federation": [
        "sso","single sign-on","saml","saml 2.0","federation","idp",
        "identity provider","conditional access","azure ad","okta","ping",
        "oidc","openid connect","oauth 2.0","identity federation","centralized auth"
    ],
    "password_policy": [
        "password","passphrase","complexity","rotation","expir","minimum length",
        "uppercase","special character","credential","password strength",
        "password reuse","password history","lockout","brute force",
        "account lockout","failed login","password reset","change password"
    ],
    # ── Access & Authorization ────────────────────────────────────────────────
    "access_control": [
        "rbac","least privilege","role-based","permission","authoriz",
        "allowlist","whitelist","ip range","just-in-time","pim","jit",
        "privilege","entitlement","access restriction","access policy",
        "access rights","user rights","group policy","access review",
        "segregation of duties","need to know","zero trust","access management",
        "geofenc","location-based","trusted location","named location",
        "conditional access","country block","network restriction",
        "trusted network","ip address restriction","login ip","ip ranges"
    ],
    "risk_based_access": [
        "risky sign","risky user","risk-based","high risk","identity protection",
        "sign-in risk","user risk","compromised","impossible travel",
        "anomalous login","suspicious login","adaptive auth","step-up",
        "continuous auth","risk policy","risk score","threat detection login",
        "risk remediat","block risky","enforce mfa risky"
    ],
    "privileged_access": [
        "privileged","admin","administrator","root","superuser","elevated",
        "privileged account","service account","break glass","emergency access",
        "standing privilege","time-bound","approval workflow","pam",
        "admin console","admin panel","elevated access"
    ],
    "user_lifecycle": [
        "provisioning","deprovisioning","scim","joiner","mover","leaver",
        "onboard","offboard","account lifecycle","user management",
        "account creation","account deletion","access revocation",
        "terminate access","disable account","role assignment","automate user"
    ],
    "session": [
        "session","timeout","idle","expiry","inactivity","auto-logout",
        "re-authentication","session duration","session management",
        "session token","persistent session","remember me","cookie"
    ],
    # ── Data Protection ───────────────────────────────────────────────────────
    "dlp": [
        "dlp","data loss prevention","sensitive data","pii","phi","pci",
        "classified","confidential","data exfiltration","information barrier",
        "personally identifiable","protected health","cardholder data",
        "sensitive information","data leakage","prevent disclosure",
        "data classification","information protection","sensitivity label"
    ],
    "external_sharing": [
        "external sharing","public link","anonymous access","guest access",
        "forwarding","share externally","data transfer","egress",
        "external user","outside organisation","third party sharing",
        "public access","open access","unrestricted sharing","outbound"
    ],
    # Split into sub-groups — at_rest and in_transit are DIFFERENT controls
    "encryption_at_rest": [
        "at rest","storage encrypt","data stored","disk encrypt","database encrypt",
        "encrypted storage","encrypt stored","encrypt database","encrypt file",
        "encrypt volume","transparent data","data at rest","rest encryption",
        "kms","key management","encrypt field","encrypt column"
    ],
    "encryption_in_transit": [
        "in transit","tls","ssl","https","transport","network encrypt",
        "data transmitted","encrypt communication","end-to-end","e2e",
        "channel encrypt","transport security","in-flight","in-transit",
        "data in motion","wire encrypt"
    ],
    "encryption_general": [
        "encrypt","cipher","aes","certificate","tokenization","encryption key",
        "cryptograph","key rotation","hsm","pkcs","x509"
    ],
    "data_residency": [
        "data residency","data sovereignty","region","country",
        "cross-border","transborder","eu","gdpr transfer","data location",
        "store data","data centre","regional","local storage",
        "geographic restriction","data geography"
    ],
    # ── Logging & Monitoring ──────────────────────────────────────────────────
    "audit_logging": [
        "audit","audit log","audit trail","unified log","log retention",
        "activity log","siem","log streaming","event log","logging",
        "log management","centralized logging","log archive","event record",
        "activity record","retain log","12 month","90 day",
        "immutable log","tamper-proof","non-repudiation"
    ],
    "monitoring_alerting": [
        "monitor","alert","anomaly","suspicious","detect","threat intel",
        "impossible travel","behavioural","unusual activity","real-time",
        "detection","notification","alarm","flag","investigate",
        "security monitoring","continuous monitoring","event correlation"
    ],
    # ── Incident & Threat ─────────────────────────────────────────────────────
    "incident_response": [
        "incident","breach","compromise","response","contain","remediat",
        "session termination","revoke","block user","incident management",
        "security incident","data breach","intrusion","isolate","quarantine",
        "recovery","restore","eradicate","post-incident"
    ],
    "threat_protection": [
        "malware","phishing","safe link","safe attachment","anti-spam",
        "defender","threat protection","sandbox","url scan","antivirus",
        "anti-malware","zero-day","ransomware","spyware","trojan",
        "email security","link protection","attachment scanning"
    ],
    # ── Integration & Apps ────────────────────────────────────────────────────
    "third_party": [
        "third-party","oauth","api access","integration","connector",
        "app permission","marketplace","connected app","admin consent",
        "external application","api key","webhook","plugin","add-on",
        "authorised app","approved app","app review","vendor access"
    ],
    # ── Infrastructure ────────────────────────────────────────────────────────
    "endpoint_device": [
        "device","endpoint","mdm","mobile","jailbreak","compliance device",
        "managed device","byod","remote wipe","device management",
        "intune","device policy","device registration","corporate device"
    ],
    "network_security": [
        "firewall","network","vpn","ip restriction","segmentation",
        "dmz","ingress","egress","private endpoint","network access",
        "network policy","subnet","vlan","network zone"
    ],
    # ── Compliance & Governance ───────────────────────────────────────────────
    "compliance_retention": [
        "compliance","retention","ediscovery","legal hold","gdpr","hipaa",
        "sox","iso 27001","nist csf","regulatory","data residency",
        "pci dss","audit requirement","regulatory requirement",
        "compliance policy","archival","preservation","discovery"
    ],
    "governance": [
        "governance","policy","procedure","standard","framework","review",
        "approval","accountability","ownership","responsibility",
        "risk management","security policy","information security"
    ],
}

KEYWORD_BOOST    = 0.08   # per matching keyword group (capped at 0.25)
DOMAIN_BOOST     = 0.05   # when control domain aligns with policy category
NONSENSE_PENALTY = 0.20   # applied when zero keyword overlap AND base score < 0.55
CONFLICT_PENALTY = 0.25   # applied when conflicting sub-groups detected

# Pairs of keyword sub-groups that CONFLICT with each other
# If one side has group A and the other exclusively has group B → penalise
CONFLICT_PAIRS = [
    ("encryption_at_rest",   "encryption_in_transit"),  # at-rest ≠ in-transit
    ("compliance_retention", "incident_response"),       # retain logs ≠ incident block
    ("user_lifecycle",       "endpoint_device"),         # user account ≠ device mgmt
    ("mfa",                  "network_security"),        # MFA auth ≠ IP/network restriction
    ("mfa",                  "data_residency"),          # MFA auth ≠ data location
    ("session",              "access_control"),          # session timeout ≠ access policy
]


def _keyword_groups_for(text: str) -> set:
    """Return set of keyword group names that match the text."""
    t = text.lower()
    return {grp for grp, kws in KEYWORD_GROUPS.items() if any(k in t for k in kws)}


def _domain_category_match(domain: str, category: str) -> bool:
    """True if control domain and policy category are semantically related."""
    mapping = {
        "access control":      ["access control","identity","governance"],
        "data protection":     ["data protection","compliance","encryption"],
        "logging & monitoring":["logging","compliance","governance"],
        "identity management": ["identity","access control","governance"],
        "incident response":   ["threat detection","threat protection","logging"],
        "compliance":          ["compliance","governance","data protection"],
        "cryptography":        ["encryption","data protection"],
        "network security":    ["network","access control"],
    }
    d = domain.lower()
    c = category.lower()
    related = mapping.get(d, [])
    return any(r in c for r in related) or any(r in d for r in [c])


def policy_encode_text(policy: Policy) -> str:
    """
    Build rich encoding text for a policy.
    1. Enrich with POLICY_NAME_LIBRARY if no/short description
    2. Expand abbreviations and concepts
    3. Repeat policy name for emphasis
    4. Append enriched description
    This solves the core problem: short OOTB policy names (e.g. "Enable SSO",
    "Trusted IP Ranges Configuration") matched against long verbose controls.
    """
    parts = []
    name = expand_concepts(expand_synonyms(policy.policy_name)) if policy.policy_name else ""
    if name:
        parts.append(name)
        parts.append(name)

    # Enrich description from library when description is missing or short
    enriched = enrich_policy_from_library(
        policy.policy_name or "",
        policy.description or ""
    )
    if enriched:
        parts.append(expand_concepts(expand_synonyms(enriched)))

    if policy.category:
        parts.append(policy.category)
    return " . ".join(parts)


def control_encode_text(control: Control) -> str:
    """
    Build rich encoding text for a control.
    1. Expand abbreviations (MFA → multi-factor authentication)
    2. Expand concepts (geofencing → conditional access location policy)
    3. Repeat control text for emphasis
    4. Append description — where detailed text lives
    """
    parts = []
    text = expand_concepts(expand_synonyms(control.control_text)) if control.control_text else ""
    if text:
        parts.append(text)
        parts.append(text)
    if control.description:
        parts.append(expand_concepts(expand_synonyms(control.description)))
    if control.subdomain:
        parts.append(control.subdomain)
    return " . ".join(parts)


def hybrid_score(base_score: float,
                 ctrl: "Control",
                 policy: "Policy") -> float:
    """
    Compute final match score combining:
      1. SecBERT cosine similarity (base)
      2. Keyword group overlap boost  — rewards shared security vocabulary
      3. Description term overlap     — rewards shared rare security terms
      4. Domain-category alignment    — rewards structural match
      5. Nonsense penalty             — suppresses irrelevant matches

    The key insight: if your description has "legacy authentication protocols
    SMTP IMAP POP3" and the policy says "Block Basic Auth SMTP IMAP POP3",
    the term overlap will boost the score even if SecBERT embedding missed it.
    """
    ctrl_text = (f"{ctrl.control_text} {ctrl.description} "
                 f"{ctrl.subdomain} {ctrl.domain}").lower()
    pol_text  = (f"{policy.policy_name} {policy.description} "
                 f"{policy.category}").lower()

    # ── 1. Keyword group overlap ──────────────────────────────────────────────
    ctrl_groups = _keyword_groups_for(ctrl_text)
    pol_groups  = _keyword_groups_for(pol_text)
    shared_groups = ctrl_groups & pol_groups
    kw_boost = min(len(shared_groups) * KEYWORD_BOOST, 0.25)

    # ── 2. Description term overlap (Jaccard on meaningful tokens) ────────────
    # Captures specific technical terms your descriptions contain
    stop = {"the","a","an","is","are","must","should","will","all","any",
            "for","to","of","in","and","or","not","be","by","with","that",
            "this","from","have","has","been","it","its","on","at","as","per"}

    def tokens(text):
        import re
        return set(re.findall(r"[a-z0-9][a-z0-9-]{2,}", text)) - stop

    ctrl_tokens = tokens(ctrl_text)
    pol_tokens  = tokens(pol_text)
    if ctrl_tokens and pol_tokens:
        intersection = ctrl_tokens & pol_tokens
        union        = ctrl_tokens | pol_tokens
        jaccard      = len(intersection) / len(union)
        # Boost proportionally — max 0.20 from description overlap
        desc_boost = min(jaccard * 2.0, 0.20)
    else:
        desc_boost = 0.0

    # ── 3. Domain-category alignment ─────────────────────────────────────────
    dom_boost = DOMAIN_BOOST if _domain_category_match(
        ctrl.domain, policy.category) else 0.0

    # ── 4. Nonsense penalty ───────────────────────────────────────────────────
    # Zero keyword group overlap → apply penalty at all score levels
    # Prevents domain_boost from falsely elevating unrelated policies
    if not shared_groups and base_score < 0.55:
        penalty = NONSENSE_PENALTY
    elif not shared_groups and base_score < 0.75:
        penalty = NONSENSE_PENALTY * 0.7   # covers mid-high base scores
    elif not shared_groups:
        penalty = NONSENSE_PENALTY * 0.3   # even high base gets small penalty
    else:
        penalty = 0.0

    # ── 5. Conflict penalty ──────────────────────────────────────────────────
    # Penalise when a known conflicting sub-group pair is detected
    # e.g. control says "at rest" but policy says "in transit" → different concepts
    conflict_penalty = 0.0
    conflict_detected = False
    for grp_a, grp_b in CONFLICT_PAIRS:
        ctrl_a = grp_a in ctrl_groups
        ctrl_b = grp_b in ctrl_groups
        pol_a  = grp_a in pol_groups
        pol_b  = grp_b in pol_groups
        # One side exclusively has A, other side exclusively has B
        if (ctrl_a and not ctrl_b) and (pol_b and not pol_a):
            conflict_penalty = max(conflict_penalty, CONFLICT_PENALTY)
            conflict_detected = True
        if (ctrl_b and not ctrl_a) and (pol_a and not pol_b):
            conflict_penalty = max(conflict_penalty, CONFLICT_PENALTY)
            conflict_detected = True

    # When conflict detected: cancel ALL boosts (they are from shared generic terms)
    # and only apply the conflict penalty on top of base score
    if conflict_detected:
        final = base_score - conflict_penalty
    else:
        final = base_score + kw_boost + desc_boost + dom_boost - penalty

    return round(min(max(final, 0.0), 1.0), 4)


# ── Domain Pairing Map ───────────────────────────────────────────────────────
# Maps each STANDARD domain → list of allowed POLICY domains.
# Controls in a standard domain will ONLY be matched against policies
# whose category is in the allowed policy domains.
# This prevents IAM controls matching against Malware Protection policies etc.
#
# Based on your domain structure from the image:
#   Standard domains  → Policy domains
#
# CUSTOMISE THIS to match your organisation's domain pairing decisions.
# Leave DOMAIN_PAIRS empty ({}) to disable filtering (match all vs all).

DOMAIN_PAIRS: dict[str, list[str]] = {
    # ── Standard Domain → Allowed Policy Domains ─────────────────────────────
    "identity and access management": [
        "mfa",
        "access control",
        "password management",
        "permissions",
        "privacy control",
        "auditing",
    ],
    "access control": [
        "access control",
        "permissions",
        "mfa",
        "password management",
        "auditing",
    ],
    "secure configuration management": [
        "secure baseline",
        "configuration and posture management",
        "permissions",
        "auditing",
        "access control",
    ],
    "data protection": [
        "data leakage protection",
        "key management",
        "privacy control",
        "auditing",
    ],
    "third-party plugin and integration management": [
        "access control",
        "permissions",
        "secure baseline",
        "auditing",
    ],
    "securityoperations": [
        "auditing",
        "malware protection",
        "operational resilience",
        "secure baseline",
    ],
    "configuration and posture management": [
        "secure baseline",
        "auditing",
        "permissions",
        "access control",
    ],
    "network services security requirements": [
        "access control",
        "secure baseline",
        "auditing",
    ],
    "governance": [
        "auditing",
        "secure baseline",
        "operational resilience",
        "privacy control",
    ],
}


def get_allowed_policy_domains(control_domain: str) -> list[str]:
    """
    Return allowed policy domains for a given control domain.
    Case-insensitive lookup.
    Returns empty list if domain not in map (meaning: allow all — no filter).
    """
    d = control_domain.strip().lower()
    # Exact match
    if d in DOMAIN_PAIRS:
        return [x.lower() for x in DOMAIN_PAIRS[d]]
    # Partial match — e.g. "Identity and Access Management" matches "identity"
    for key, allowed in DOMAIN_PAIRS.items():
        if key in d or d in key:
            return [x.lower() for x in allowed]
    return []   # no match → no filter applied (allow all)


# Tracks policy categories seen that are NOT in any domain pair
# Populated during run — printed as warning so user can add them
_UNKNOWN_POLICY_DOMAINS: set = set()


def is_policy_domain_allowed(control_domain: str, policy_category: str) -> bool:
    """
    Return True if policy_category is allowed for this control_domain.

    Rules:
      - Domain filtering disabled → always True
      - Control domain not in DOMAIN_PAIRS → always True (allow all for that control)
      - Policy category not in ANY domain pair anywhere → True + warning collected
        (new domain discovered — should be added to domain_pairs.json)
      - Policy category in allowed list for this control → True
      - Otherwise → False (blocked)
    """
    if not DOMAIN_PAIRS:
        return True   # domain filtering disabled

    allowed = get_allowed_policy_domains(control_domain)
    if not allowed:
        return True   # no mapping defined for this control domain → allow all

    pol_cat = policy_category.strip().lower()

    # Check if this policy category exists ANYWHERE in the domain pairs
    # If not found anywhere → it is a NEW/UNKNOWN domain → allow + track
    all_known_domains = {
        d.lower()
        for domains in DOMAIN_PAIRS.values()
        for d in domains
    }
    is_known = any(
        k in pol_cat or pol_cat in k
        for k in all_known_domains
    )
    if not is_known and pol_cat:
        _UNKNOWN_POLICY_DOMAINS.add(policy_category.strip())
        return True   # unknown domain → allow through, user should add to pairs

    return any(a in pol_cat or pol_cat in a for a in allowed)


# ── Core Mapper ───────────────────────────────────────────────────────────────

class SSPMMapper:
    """
    Maps org security controls to OOTB policies for ONE SaaS app at a time.

    Workflow:
        mapper = SSPMMapper()
        mapper.load_controls("controls.csv")
        mapper.load_policies("policies.csv", app="Microsoft 365")
        report = mapper.run()
        mapper.print_report(report)
        mapper.save_report(report, "output.json")
    """

    def __init__(self, model_path: str = None):
        self.encoder   = SecBERTEncoder(model_path=model_path)
        self.cache     = EmbeddingCache()
        self.controls: list[Control] = []
        self.policies: list[Policy]  = []
        self.app_name: str = ""

    # ── Loaders ───────────────────────────────────────────────────────────────

    def load_controls(self, path: str):
        self.controls = ControlLoader.load(path)
        domains = sorted(set(c.domain for c in self.controls))
        print(f"📋  Controls loaded : {len(self.controls)} controls  [from: {path}]")
        print(f"     Domains        : {', '.join(domains)}")
        print(f"     Frameworks     : {', '.join(sorted(set(c.framework for c in self.controls if c.framework)))}")
        # Show first 3 IDs so user can confirm IDs are read from their file
        sample_ids = [c.control_id for c in self.controls[:3]]
        print(f"     Sample IDs     : {sample_ids}  ← from your file")
        print()

    def load_controls_from_list(self, controls: list[Control]):
        self.controls = controls
        print(f"📋  Controls loaded : {len(self.controls)}\n")

    def load_policies(self, path: str, app: str = ""):
        """
        Load OOTB policies from a CSV or JSON file.

        app parameter is used as a label only (stored in report).
        If your CSV has an app_name column, rows are filtered to match app.
        If your CSV has NO app_name column (single-app file), all rows are loaded.
        """
        self.app_name = app
        # Always try loading all rows first for single-app files
        self.policies = PolicyLoader.load(path, app_filter=None)
        # If multi-app file AND app specified, filter down
        if app and self.policies:
            has_app_col = any(
                hasattr(p, "_app_col") or
                True  # we rely on from_csv internal detection
                for p in self.policies[:1]
            )
            # Check if CSV actually had an app_name column by re-examining
            filtered = PolicyLoader.load(path, app_filter=app)
            if len(filtered) > 0:
                self.policies = filtered
                print(f"📂  Policies loaded : {len(self.policies)} policies  [app: {app}  (filtered from {path})]")
                sample_ids = [p.policy_id for p in self.policies[:3]]
                print(f"     Sample IDs     : {sample_ids}  ← from your file")
            else:
                # No app_name column — single-app file, keep all rows
                print(f"📂  Policies loaded : {len(self.policies)} policies  [app: {app}  (single-app file: {path})]")
                sample_ids = [p.policy_id for p in self.policies[:3]]
                print(f"     Sample IDs     : {sample_ids}  ← from your file")
        else:
            print(f"📂  Policies loaded : {len(self.policies)} policies  [app: {app or 'all'}  from: {path}]")
            sample_ids = [p.policy_id for p in self.policies[:3]]
            print(f"     Sample IDs     : {sample_ids}  ← from your file")
        if len(self.policies) == 0:
            print(f"⚠️   0 policies loaded — check your file path and column names.")
            return
        cats = sorted(set(p.category for p in self.policies if p.category))
        if cats:
            print(f"     Categories     : {', '.join(cats)}\n")

    def load_policies_from_list(self, policies: list[Policy], app: str = ""):
        self.policies = policies
        self.app_name = app

    # ── Run Mapping ───────────────────────────────────────────────────────────

    def run(self,
            top_k:     int   = 5,
            threshold: float = THRESHOLD_MIN) -> dict:
        """
        Run the full bidirectional mapping and return a structured report.
        """
        if not self.controls:
            raise RuntimeError("No controls loaded. Call load_controls() first.")
        if not self.policies:
            raise RuntimeError("No policies loaded. Call load_policies() first.")

        import time as _time
        t_total = _time.time()
        print(f"🔄  Encoding {len(self.controls)} controls and {len(self.policies)} policies...")

        # Build encode texts
        ctrl_texts   = [control_encode_text(c) for c in self.controls]
        policy_texts = [policy_encode_text(p)  for p in self.policies]

        # Fit TF-IDF if needed
        if self.encoder.mode == "tfidf":
            self.encoder.fit_tfidf(ctrl_texts + policy_texts)

        # ── Encode policies — cached per app ─────────────────────────────────
        cache_label = self.app_name or "policies"
        policy_vecs = self.cache.get(cache_label, policy_texts)
        if policy_vecs is None:
            print(f"     Encoding {len(policy_texts)} policies in batches...")
            t0 = __import__("time").time()
            policy_vecs = self.encoder.encode(policy_texts)
            elapsed = __import__("time").time() - t0
            self.cache.set(cache_label, policy_texts, policy_vecs)
            print(f"     Policy vectors encoded in {elapsed:.1f}s and cached.")
        else:
            print(f"     Policy vectors loaded from cache (instant).")

        # ── Encode controls — also cached ─────────────────────────────────────
        # Controls change less often than you might think — cache them too
        ctrl_cache_label = f"controls__{self.app_name or 'default'}"
        ctrl_vecs = self.cache.get(ctrl_cache_label, ctrl_texts)
        if ctrl_vecs is None:
            print(f"     Encoding {len(ctrl_texts)} controls in batches...")
            t0 = __import__("time").time()
            ctrl_vecs = self.encoder.encode(ctrl_texts)
            elapsed = __import__("time").time() - t0
            self.cache.set(ctrl_cache_label, ctrl_texts, ctrl_vecs)
            print(f"     Control vectors encoded in {elapsed:.1f}s and cached.")
        else:
            print(f"     Control vectors loaded from cache (instant).")

        t_enc = _time.time() - t_total
        print(f"✅  Encoding complete in {t_enc:.1f}s\n")

        # ── Build similarity matrix ───────────────────────────────────────────
        # Shape: (n_controls, n_policies)
        ctrl_vecs   = np.atleast_2d(ctrl_vecs).astype(np.float32)   # float32 = 2x faster
        policy_vecs = np.atleast_2d(policy_vecs).astype(np.float32)

        if ctrl_vecs.shape[1] != policy_vecs.shape[1]:
            raise RuntimeError(
                f"Embedding dimension mismatch: controls={ctrl_vecs.shape[1]} "
                f"vs policies={policy_vecs.shape[1]}. "
                f"Clear policy_cache/ and re-run: rm -rf policy_cache/"
            )

        # Normalise rows to unit length — dot product then equals cosine similarity
        def norm(vecs: np.ndarray) -> np.ndarray:
            n = np.linalg.norm(vecs, axis=1, keepdims=True)
            n[n == 0] = 1e-9          # avoid division by zero
            return vecs / n

        # Single matrix multiply gives all N×M similarities at once
        sim_matrix = norm(ctrl_vecs) @ norm(policy_vecs).T
        # sim_matrix[i][j] = cosine similarity between control i and policy j

        # ── Build domain filter mask ──────────────────────────────────────────
        # For each (control i, policy j) pair, check if the policy domain
        # is allowed for this control's standard domain.
        # Disallowed pairs are masked to 0.0 before scoring.
        n_masked = 0
        if DOMAIN_PAIRS:
            domain_mask = np.ones((len(self.controls), len(self.policies)),
                                   dtype=np.float32)
            for i, ctrl in enumerate(self.controls):
                for j, pol in enumerate(self.policies):
                    if not is_policy_domain_allowed(ctrl.domain, pol.category):
                        domain_mask[i, j] = 0.0
                        n_masked += 1
            sim_matrix = sim_matrix * domain_mask
            pct = round(n_masked / (len(self.controls) * len(self.policies)) * 100)
            print(f"     Domain filter: {n_masked} pairs masked ({pct}% of matrix) "
                  f"— only allowed domain pairs compared.")
            # Warn about any new policy domains not yet in domain_pairs.json
            if _UNKNOWN_POLICY_DOMAINS:
                print(f"\n  ⚠️   NEW POLICY DOMAINS DETECTED — not in domain_pairs.json:")
                for ud in sorted(_UNKNOWN_POLICY_DOMAINS):
                    print(f"       → '{ud}'  (allowed through — add to domain_pairs.json to control pairing)")
                print()
        else:
            print(f"     Domain filter: disabled — all controls vs all policies.")

        # ── Forward pass: control → policies ─────────────────────────────────
        control_results: list[ControlResult] = []
        matched_policy_ids = set()

        for i, ctrl in enumerate(self.controls):
            raw_sims = sim_matrix[i]

            # ── Apply hybrid scoring ──────────────────────────────────────────
            boosted_sims = np.array([
                hybrid_score(float(raw_sims[j]), ctrl, self.policies[j])
                if raw_sims[j] > 0.0   # skip masked pairs entirely
                else 0.0
                for j in range(len(self.policies))
            ])

            ranked  = np.argsort(boosted_sims)[::-1]
            matches = []

            for j in ranked[:top_k]:
                score     = float(boosted_sims[j])
                raw_score = float(raw_sims[j])

                # Dynamic threshold: lower it when keyword overlap is strong
                # This ensures keyword-matched policies are never dropped
                ctrl_kw_set = _keyword_groups_for(
                    f"{ctrl.control_text} {ctrl.description}")
                pol_kw_set  = _keyword_groups_for(
                    f"{self.policies[j].policy_name} "
                    f"{self.policies[j].description} "
                    f"{self.policies[j].category}")
                shared_count = len(ctrl_kw_set & pol_kw_set)
                effective_threshold = (
                    0.15 if shared_count >= 2 else   # strong keyword overlap
                    0.25 if shared_count == 1 else   # some keyword overlap
                    threshold                          # no keyword overlap — use default
                )

                if score < effective_threshold:
                    break
                coverage = (
                    "FULL"     if score >= THRESHOLD_FULL     else
                    "PARTIAL"  if score >= THRESHOLD_PARTIAL  else
                    "INDIRECT"
                )
                pol = self.policies[j]
                # Confidence = how well keyword groups align (independent of score)
                ctrl_text = f"{ctrl.control_text} {ctrl.description}"
                pol_text  = f"{pol.policy_name} {pol.description} {pol.category}"
                shared_kw = _keyword_groups_for(ctrl_text) & _keyword_groups_for(pol_text)
                confidence = ("HIGH"   if len(shared_kw) >= 2 else
                              "MEDIUM" if len(shared_kw) == 1 else
                              "LOW")

                matches.append(PolicyMatch(
                    policy_id        = pol.policy_id,
                    policy_name      = pol.policy_name,
                    policy_category  = pol.category,
                    similarity_score = score,
                    coverage         = coverage,
                    impact           = pol.impact,
                    confidence       = confidence,
                ))
                matched_policy_ids.add(pol.policy_id)

            is_covered = any(m.coverage in ("FULL", "PARTIAL", "INDIRECT") for m in matches)

            control_results.append(ControlResult(
                control_id   = ctrl.control_id,
                control_text = ctrl.control_text,
                domain       = ctrl.domain,
                framework    = ctrl.framework,
                subdomain    = ctrl.subdomain,
                risk_level   = self._assess_risk(ctrl.control_text),
                matches      = matches,
                is_covered   = is_covered,
            ))

        # ── Reverse pass: policy → controls ──────────────────────────────────
        policy_results: list[PolicyResult] = []

        for j, pol in enumerate(self.policies):
            sims = sim_matrix[:, j]   # all controls vs this policy
            matched_ctrl_ids = []

            for i, ctrl in enumerate(self.controls):
                score = float(sims[i])
                if score >= THRESHOLD_INDIRECT:
                    matched_ctrl_ids.append(ctrl.control_id)

            is_orphan = len(matched_ctrl_ids) == 0
            policy_results.append(PolicyResult(
                policy_id        = pol.policy_id,
                policy_name      = pol.policy_name,
                policy_category  = pol.category,
                matched_controls = matched_ctrl_ids,
                is_orphan        = is_orphan,
            ))

        # ── Relationship classification ───────────────────────────────────────
        relationships = self._classify_relationships(control_results, policy_results)

        # ── Domain summary ────────────────────────────────────────────────────
        domain_summary = self._domain_summary(control_results)

        # ── Build final report ────────────────────────────────────────────────
        uncovered_controls = [r for r in control_results if not r.is_covered]
        orphan_policies    = [r for r in policy_results  if r.is_orphan]

        report = {
            "app":              self.app_name or "Unknown App",
            "encoder_mode":     self.encoder.mode,
            "unknown_policy_domains": sorted(_UNKNOWN_POLICY_DOMAINS),
            "summary": {
                "total_controls":       len(self.controls),
                "total_policies":       len(self.policies),
                "covered_controls":     len([r for r in control_results if r.is_covered]),
                "uncovered_controls":   len(uncovered_controls),
                "orphan_policies":      len(orphan_policies),
                "full_matches":         sum(1 for r in control_results
                                           for m in r.matches if m.coverage == "FULL"),
                "partial_matches":      sum(1 for r in control_results
                                           for m in r.matches if m.coverage == "PARTIAL"),
                "indirect_matches":     sum(1 for r in control_results
                                           for m in r.matches if m.coverage == "INDIRECT"),
            },
            "domain_summary":   domain_summary,
            "relationships":    relationships,
            "control_mappings": [self._serialize_ctrl(r) for r in control_results],
            "policy_mappings":  [self._serialize_pol(r)  for r in policy_results],
            "uncovered_controls": [
                {
                    "control_id":   r.control_id,
                    "control_text": r.control_text,
                    "domain":       r.domain,
                    "framework":    r.framework,
                    "subdomain":    r.subdomain,
                    "risk_level":   r.risk_level,
                    "reason":       "No OOTB policy matches at or above threshold",
                }
                for r in uncovered_controls
            ],
            "orphan_policies": [
                {
                    "policy_id":       r.policy_id,
                    "policy_name":     r.policy_name,
                    "policy_category": r.policy_category,
                    "reason":          "No org control maps to this OOTB policy",
                }
                for r in orphan_policies
            ],
        }
        return report

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _classify_relationships(self,
                                 ctrl_results: list[ControlResult],
                                 pol_results:  list[PolicyResult]) -> dict:
        """
        Classify all mapping relationships into:
          one_to_one    : 1 control  → exactly 1 policy
          one_to_many   : 1 control  → 2+ policies
          many_to_one   : 2+ controls → same 1 policy
          many_to_many  : 2+ controls share 2+ overlapping policies

        Includes ALL match types: FULL, PARTIAL, INDIRECT.
        """
        # ── Build control → policies map (ALL coverage types) ────────────────
        ctrl_to_pols = {}
        for r in ctrl_results:
            pol_ids = set(
                m.policy_id for m in r.matches
                if m.coverage in ("FULL", "PARTIAL", "INDIRECT")
            )
            if pol_ids:
                ctrl_to_pols[r.control_id] = pol_ids

        # ── Build policy → controls map (ALL coverage types) ─────────────────
        # Re-derive from ctrl_to_pols so it's consistent (not from reverse pass
        # which used a different threshold)
        pol_to_ctrls = {}
        for ctrl_id, pol_ids in ctrl_to_pols.items():
            for pid in pol_ids:
                pol_to_ctrls.setdefault(pid, set()).add(ctrl_id)

        one_to_one   = []
        one_to_many  = []
        many_to_one  = []
        many_to_many = []
        seen_m2m_clusters = set()
        # Track policies already classified as many_to_one to avoid duplicates
        seen_m2o_policies = set()

        for ctrl_id, pol_ids in ctrl_to_pols.items():
            n_pols = len(pol_ids)

            if n_pols == 1:
                # ── This control maps to exactly 1 policy ─────────────────────
                pid = next(iter(pol_ids))
                n_ctrls_for_this_pol = len(pol_to_ctrls.get(pid, set()))

                if n_ctrls_for_this_pol == 1:
                    # 1 control ↔ 1 policy
                    one_to_one.append({
                        "control": ctrl_id,
                        "policy":  pid,
                    })
                else:
                    # Multiple controls → same 1 policy (many-to-1)
                    if pid not in seen_m2o_policies:
                        seen_m2o_policies.add(pid)
                        many_to_one.append({
                            "policy":   pid,
                            "controls": sorted(pol_to_ctrls[pid]),
                        })

            else:
                # ── This control maps to 2+ policies ─────────────────────────
                # Check if any of those policies are ALSO hit by other controls
                other_controls_exist = any(
                    len(pol_to_ctrls.get(pid, set())) > 1
                    for pid in pol_ids
                )

                if other_controls_exist:
                    # Overlapping cluster → many-to-many
                    # Expand cluster to its maximal connected set
                    cluster_ctrls = set([ctrl_id])
                    cluster_pols  = set(pol_ids)
                    # Keep expanding until stable
                    prev_size = 0
                    while prev_size != len(cluster_ctrls) + len(cluster_pols):
                        prev_size = len(cluster_ctrls) + len(cluster_pols)
                        for pid in list(cluster_pols):
                            cluster_ctrls.update(pol_to_ctrls.get(pid, set()))
                        for cid in list(cluster_ctrls):
                            cluster_pols.update(ctrl_to_pols.get(cid, set()))

                    # Use frozenset of controls as dedup key
                    cluster_key = frozenset(cluster_ctrls)
                    if cluster_key not in seen_m2m_clusters:
                        seen_m2m_clusters.add(cluster_key)
                        many_to_many.append({
                            "controls": sorted(cluster_ctrls),
                            "policies": sorted(cluster_pols),
                        })
                else:
                    # Only this control hits these policies → genuine 1-to-many
                    one_to_many.append({
                        "control":  ctrl_id,
                        "policies": sorted(pol_ids),
                    })

        return {
            "one_to_one":   one_to_one,
            "one_to_many":  one_to_many,
            "many_to_one":  many_to_one,
            "many_to_many": many_to_many,
        }

    def _domain_summary(self, ctrl_results: list[ControlResult]) -> list[dict]:
        """Summarise coverage per security domain."""
        from collections import defaultdict
        domains = defaultdict(lambda: {"total": 0, "covered": 0,
                                        "uncovered": [], "high_risk": 0})
        for r in ctrl_results:
            d = domains[r.domain]
            d["total"] += 1
            # Covered = any match (FULL, PARTIAL, or INDIRECT)
            if r.is_covered:
                d["covered"] += 1
            else:
                d["uncovered"].append(r.control_id)
            # Track coverage breakdown
            for m in r.matches:
                d[m["coverage"] if isinstance(m, dict) else m.coverage] =                     d.get(m["coverage"] if isinstance(m, dict) else m.coverage, 0) + 1
            if r.risk_level == "HIGH":
                d["high_risk"] += 1

        summary = []
        for domain, stats in sorted(domains.items()):
            total    = stats["total"]
            covered  = stats["covered"]
            pct      = round(covered / total * 100) if total else 0
            n_full     = stats.get("FULL", 0)
            n_partial  = stats.get("PARTIAL", 0)
            n_indirect = stats.get("INDIRECT", 0)
            summary.append({
                "domain":               domain,
                "total_controls":       total,
                "covered_controls":     covered,
                "uncovered_controls":   total - covered,
                "coverage_pct":         pct,
                "full_matches":         n_full,
                "partial_matches":      n_partial,
                "indirect_matches":     n_indirect,
                "high_risk_controls":   stats["high_risk"],
                "uncovered_control_ids": stats["uncovered"],
                "coverage_status":     (
                    "FULL"    if pct == 100 else
                    "GOOD"    if pct >= 75  else
                    "PARTIAL" if pct >= 50  else
                    "POOR"
                ),
                "coverage_note": (
                    "Includes INDIRECT matches in coverage %" 
                    if n_indirect > 0 else ""
                ),
            })
        return summary

    def explain_match(self, control_id: str, policy_id: str):
        """
        Print detailed breakdown of WHY a control-policy pair scored what it did.
        Use this to understand and debug bad/missing matches.

        Usage:
            mapper.explain_match("ISO-AC-001", "M365-POL-001")
        """
        ctrl = next((c for c in self.controls if c.control_id == control_id), None)
        pol  = next((p for p in self.policies  if p.policy_id  == policy_id),  None)

        if not ctrl:
            print(f"Control '{control_id}' not found.")
            return
        if not pol:
            print(f"Policy '{policy_id}' not found.")
            return

        ctrl_enc = control_encode_text(ctrl)
        pol_enc  = policy_encode_text(pol)

        ctrl_full = f"{ctrl.control_text} {ctrl.description} {ctrl.subdomain} {ctrl.domain}".lower()
        pol_full  = f"{pol.policy_name} {pol.description} {pol.category}".lower()

        ctrl_groups = _keyword_groups_for(ctrl_full)
        pol_groups  = _keyword_groups_for(pol_full)
        shared      = ctrl_groups & pol_groups

        import re
        stop = {"the","a","an","is","are","must","should","will","all","any",
                "for","to","of","in","and","or","not","be","by","with","that",
                "this","from","have","has","been","it","its","on","at","as","per"}
        def tokens(t): return set(re.findall(r"[a-z0-9][a-z0-9-]{2,}", t)) - stop
        ctrl_tok = tokens(ctrl_full)
        pol_tok  = tokens(pol_full)
        shared_tok = ctrl_tok & pol_tok
        jaccard = len(shared_tok) / len(ctrl_tok | pol_tok) if (ctrl_tok | pol_tok) else 0

        dom_match = _domain_category_match(ctrl.domain, pol.category)

        print(f"\n{'='*60}")
        print(f"  MATCH EXPLANATION")
        print(f"{'='*60}")
        print(f"  Control  : [{ctrl.control_id}] {ctrl.control_text[:80]}")
        print(f"  Policy   : [{pol.policy_id}] {pol.policy_name}")
        print(f"{'─'*60}")
        print(f"  Encoded control text:")
        print(f"    {ctrl_enc[:200]}")
        print(f"  Encoded policy text:")
        print(f"    {pol_enc[:200]}")
        print(f"{'─'*60}")
        print(f"  Keyword groups — Control  : {sorted(ctrl_groups)}")
        print(f"  Keyword groups — Policy   : {sorted(pol_groups)}")
        print(f"  Shared groups             : {sorted(shared)}  →  boost = +{min(len(shared)*0.10,0.30):.2f}")
        print(f"  Shared tokens             : {sorted(list(shared_tok))[:15]}")
        print(f"  Jaccard token overlap     : {jaccard:.4f}  →  boost = +{min(jaccard*2,0.20):.2f}")
        print(f"  Domain-category match     : {dom_match}  →  boost = +{0.06 if dom_match else 0:.2f}")
        total_boost = min(len(shared)*0.10,0.30) + min(jaccard*2,0.20) + (0.06 if dom_match else 0)
        print(f"  Total boost               : +{total_boost:.2f}")
        print(f"{'─'*60}")
        print(f"  To get base SecBERT score : call mapper.run() first then check")
        print(f"  Final score ≈ base_cosine + {total_boost:.2f} (minus penalty if base < 0.5)")
        print(f"{'='*60}\n")

    def _assess_risk(self, text: str) -> str:
        t = text.lower()
        high = sum(1 for k in [
            "admin","privilege","mfa","multi-factor","password","encrypt",
            "credential","authentication","pii","phi","audit","gdpr","hipaa","root"
        ] if k in t)
        med = sum(1 for k in [
            "log","monitor","share","external","third-party","session","backup","patch"
        ] if k in t)
        return "HIGH" if high >= 2 else "MEDIUM" if (high == 1 or med >= 2) else "LOW"

    def _serialize_ctrl(self, r: ControlResult) -> dict:
        return {
            "control_id":   r.control_id,
            "control_text": r.control_text,
            "domain":       r.domain,
            "framework":    r.framework,
            "subdomain":    r.subdomain,
            "risk_level":   r.risk_level,
            "is_covered":   r.is_covered,
            "match_count":  len(r.matches),
            "matches": [
                {
                    "policy_id":        m.policy_id,
                    "policy_name":      m.policy_name,
                    "policy_category":  m.policy_category,
                    "similarity_score": m.similarity_score,
                    "coverage":         m.coverage,
                    "impact":           m.impact,
                    "confidence":       m.confidence,
                }
                for m in r.matches
            ],
        }

    def _serialize_pol(self, r: PolicyResult) -> dict:
        return {
            "policy_id":         r.policy_id,
            "policy_name":       r.policy_name,
            "policy_category":   r.policy_category,
            "matched_controls":  r.matched_controls,
            "match_count":       len(r.matched_controls),
            "is_orphan":         r.is_orphan,
        }

    # ── Report Printer ────────────────────────────────────────────────────────

    def print_report(self, report: dict, verbose: bool = False):
        R="\033[0m"; B="\033[1m"
        RED="\033[91m"; YEL="\033[93m"; GRN="\033[92m"
        CYN="\033[96m"; BLU="\033[94m"; DIM="\033[2m"; MAG="\033[95m"

        s = report["summary"]
        print(f"\n{'═'*70}")
        print(f"{B}{CYN}  SSPM MAPPING REPORT  —  {report['app']}{R}")
        print(f"{'═'*70}")
        print(f"  Encoder     : {DIM}{report['encoder_mode']}{R}")
        print(f"  Controls    : {s['total_controls']}   Policies : {s['total_policies']}")
        print(f"  Covered     : {GRN}{s['covered_controls']}{R}   "
              f"Uncovered : {RED}{s['uncovered_controls']}{R}   "
              f"Orphan policies : {YEL}{s['orphan_policies']}{R}")
        print(f"  FULL        : {GRN}{s['full_matches']}{R}   "
              f"PARTIAL : {YEL}{s['partial_matches']}{R}   "
              f"INDIRECT : {DIM}{s['indirect_matches']}{R}   "
              f"{DIM}(all 3 count toward coverage %){R}")

        # ── Domain Summary ────────────────────────────────────────────────────
        print(f"\n{'─'*70}")
        print(f"  {B}{BLU}DOMAIN COVERAGE SUMMARY{R}")
        print(f"{'─'*70}")
        sc = {"FULL":GRN, "GOOD":GRN, "PARTIAL":YEL, "POOR":RED}
        for d in report["domain_summary"]:
            bar_len = int(d["coverage_pct"] / 5)
            bar = "█"*bar_len + "░"*(20-bar_len)
            col = sc.get(d["coverage_status"], R)
            hr  = f"  {RED}⚠ {d['high_risk_controls']} HIGH{R}" if d["high_risk_controls"] else ""
            # Show breakdown: F=FULL P=PARTIAL I=INDIRECT
            breakdown = (f"{GRN}F:{d.get('full_matches',0)}{R} "
                         f"{YEL}P:{d.get('partial_matches',0)}{R} "
                         f"{DIM}I:{d.get('indirect_matches',0)}{R}")
            print(f"  {B}{d['domain']:<28}{R} "
                  f"{col}{bar}{R} {d['coverage_pct']:>3}%  "
                  f"({d['covered_controls']}/{d['total_controls']})  "
                  f"{breakdown}{hr}")

        # ── Relationship types ────────────────────────────────────────────────
        rel = report["relationships"]
        print(f"\n{'─'*70}")
        print(f"  {B}{MAG}MAPPING RELATIONSHIPS{R}")
        print(f"{'─'*70}")
        print(f"  1-to-1   : {len(rel['one_to_one']):<4} "
              f"  1-to-many : {len(rel['one_to_many']):<4} "
              f"  many-to-1 : {len(rel['many_to_one']):<4} "
              f"  many-to-many : {len(rel['many_to_many'])}")

        # ── Control Mappings ──────────────────────────────────────────────────
        if verbose:
            print(f"\n{'─'*70}")
            print(f"  {B}{CYN}CONTROL → POLICY MAPPINGS{R}")
            print(f"{'─'*70}")
            for cm in report["control_mappings"]:
                cov_icon = f"{GRN}✔{R}" if cm["is_covered"] else f"{RED}✘{R}"
                rc = {"HIGH":RED,"MEDIUM":YEL,"LOW":GRN}.get(cm["risk_level"], R)
                print(f"\n  {cov_icon} [{cm['control_id']}] {rc}{cm['risk_level']}{R}  "
                      f"{DIM}{cm['domain']}{R}  {cm['framework']}")
                print(f"     {cm['control_text'][:90]}")
                if cm["matches"]:
                    for m in cm["matches"]:
                        cc = {"FULL":GRN,"PARTIAL":YEL,"INDIRECT":DIM}.get(m["coverage"],R)
                        cat = f" [{m['policy_category']}]" if m["policy_category"] else ""
                        print(f"       {cc}→ {m['policy_name']}{cat}  "
                              f"({m['similarity_score']:.3f} {m['coverage']}){R}")
                else:
                    print(f"       {RED}→ No matches{R}")

        # ── Uncovered Controls ────────────────────────────────────────────────
        print(f"\n{'─'*70}")
        print(f"  {B}{RED}UNCOVERED CONTROLS "
              f"({len(report['uncovered_controls'])}) "
              f"— Standards with no matching OOTB policy{R}")
        print(f"{'─'*70}")
        if not report["uncovered_controls"]:
            print(f"  {GRN}✔ All controls have at least one matching policy.{R}")
        else:
            for uc in report["uncovered_controls"]:
                rc = {"HIGH":RED,"MEDIUM":YEL,"LOW":GRN}.get(uc["risk_level"],R)
                sub = f" / {uc['subdomain']}" if uc["subdomain"] else ""
                print(f"  {RED}✘{R} [{uc['control_id']}] "
                      f"{rc}{uc['risk_level']}{R}  "
                      f"{DIM}{uc['domain']}{sub}  {uc['framework']}{R}")
                print(f"     {uc['control_text'][:100]}")

        # ── Orphan Policies ───────────────────────────────────────────────────
        print(f"\n{'─'*70}")
        print(f"  {B}{YEL}ORPHAN POLICIES "
              f"({len(report['orphan_policies'])}) "
              f"— OOTB policies not required by any standard{R}")
        print(f"{'─'*70}")
        if not report["orphan_policies"]:
            print(f"  {GRN}✔ All policies are referenced by at least one control.{R}")
        else:
            for op in report["orphan_policies"]:
                cat = f" [{op['policy_category']}]" if op["policy_category"] else ""
                print(f"  {YEL}○{R} [{op['policy_id']}] {op['policy_name']}{DIM}{cat}{R}")

        print(f"\n{'═'*70}\n")

    def save_report(self, report: dict, path: str = "mapping_report.json"):
        with open(path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2)
        print(f"📁  Report saved to {path}")

    def save_xlsx(self, report: dict, path: str = "mapping_report.xlsx"):
        """
        Export all mapping results to a single Excel workbook with 6 sheets:
          1. Summary          — overall counts + domain coverage table
          2. Control Mappings — every control with all matched policies (one row per pair)
          3. Policy Mappings  — every policy with all matched controls (one row per pair)
          4. Uncovered Controls — controls with NO matching OOTB policy
          5. Orphan Policies  — policies with NO matching control
          6. Relationships    — 1:1, 1:N, N:1, N:N mapping types
        """
        from openpyxl import Workbook
        from openpyxl.styles import (Font, PatternFill, Alignment,
                                      Border, Side, GradientFill)
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        app = report.get("app", "")

        # ── Colour palette ────────────────────────────────────────────────────
        C_HEADER_DARK   = "1F3864"   # dark navy  — sheet headers
        C_HEADER_MID    = "2E75B6"   # mid blue   — section headers
        C_HEADER_LIGHT  = "D6E4F0"   # light blue — column headers
        C_FULL          = "C6EFCE"   # green bg   — FULL coverage
        C_FULL_FG       = "276221"
        C_PARTIAL       = "FFEB9C"   # amber bg   — PARTIAL
        C_PARTIAL_FG    = "9C5700"
        C_INDIRECT      = "EDEDED"   # grey bg    — INDIRECT
        C_INDIRECT_FG   = "595959"
        C_UNCOVERED     = "FCE4D6"   # red-ish bg — uncovered / orphan
        C_UNCOVERED_FG  = "843C0C"
        C_HIGH          = "FFC7CE"   # risk HIGH
        C_HIGH_FG       = "9C0006"
        C_MEDIUM        = "FFEB9C"
        C_MEDIUM_FG     = "9C5700"
        C_LOW           = "C6EFCE"
        C_LOW_FG        = "276221"
        C_ALT_ROW       = "F2F7FC"   # alternating row tint

        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def hdr_font(size=11, bold=True, color="FFFFFF"):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def body_font(size=10, bold=False, color="000000"):
            return Font(name="Arial", size=size, bold=bold, color=color)

        def fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def wrap_align(horizontal="left"):
            return Alignment(horizontal=horizontal, vertical="top",
                             wrap_text=True)

        def set_col_widths(ws, widths: dict):
            for col_letter, w in widths.items():
                ws.column_dimensions[col_letter].width = w

        def freeze(ws, cell="A2"):
            ws.freeze_panes = cell

        def write_header_row(ws, row_num, headers, bg=C_HEADER_LIGHT, fg="1F3864"):
            for col, h in enumerate(headers, 1):
                c = ws.cell(row=row_num, column=col, value=h)
                c.font      = hdr_font(size=10, color=fg)
                c.fill      = fill(bg)
                c.alignment = Alignment(horizontal="center", vertical="center",
                                        wrap_text=True)
                c.border    = border

        def write_data_row(ws, row_num, values, alt=False):
            bg = C_ALT_ROW if alt else "FFFFFF"
            for col, v in enumerate(values, 1):
                c = ws.cell(row=row_num, column=col, value=v)
                c.font      = body_font()
                c.fill      = fill(bg)
                c.alignment = wrap_align()
                c.border    = border

        def color_cell(cell, bg, fg):
            cell.fill = fill(bg)
            cell.font = Font(name="Arial", size=10, bold=True, color=fg)

        def title_row(ws, text, ncols, row=1):
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=ncols)
            c = ws.cell(row=row, column=1, value=text)
            c.font      = Font(name="Arial", size=13, bold=True, color="FFFFFF")
            c.fill      = fill(C_HEADER_DARK)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row].height = 28

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 1 — Summary
        # ═══════════════════════════════════════════════════════════════════════
        ws1 = wb.active
        ws1.title = "Summary"
        s = report["summary"]

        title_row(ws1, f"SSPM Mapping Report  —  {app}", 6)

        # KPI block
        kpis = [
            ("Total Controls",     s["total_controls"]),
            ("Total Policies",     s["total_policies"]),
            ("Covered Controls",   s["covered_controls"]),
            ("Uncovered Controls", s["uncovered_controls"]),
            ("Orphan Policies",    s["orphan_policies"]),
            ("FULL Matches",       s["full_matches"]),
            ("PARTIAL Matches",    s["partial_matches"]),
            ("INDIRECT Matches",   s["indirect_matches"]),
        ]
        ws1.cell(row=3, column=1, value="Metric").font  = hdr_font(color="1F3864")
        ws1.cell(row=3, column=2, value="Value").font   = hdr_font(color="1F3864")
        ws1.cell(row=3, column=1).fill = fill(C_HEADER_LIGHT)
        ws1.cell(row=3, column=2).fill = fill(C_HEADER_LIGHT)
        for i, (label, val) in enumerate(kpis, 4):
            ws1.cell(row=i, column=1, value=label).font  = body_font(bold=True)
            ws1.cell(row=i, column=2, value=val).font    = body_font()
            ws1.cell(row=i, column=1).fill = fill(C_ALT_ROW if i % 2 == 0 else "FFFFFF")
            ws1.cell(row=i, column=2).fill = fill(C_ALT_ROW if i % 2 == 0 else "FFFFFF")
            ws1.cell(row=i, column=1).border = border
            ws1.cell(row=i, column=2).border = border

        # Domain summary table
        start_row = 14
        ws1.cell(row=start_row, column=1,
                 value="Domain Coverage  (Coverage % includes FULL + PARTIAL + INDIRECT)").font = hdr_font(size=11, color="1F3864")
        ws1.cell(row=start_row, column=1).fill = fill(C_HEADER_LIGHT)

        hdr = ["Domain", "Total", "Covered", "Uncovered",
               "Coverage %", "FULL", "PARTIAL", "INDIRECT", "Status", "Note"]
        write_header_row(ws1, start_row + 1, hdr)
        sc_map = {"FULL":"C6EFCE","GOOD":"C6EFCE","PARTIAL":"FFEB9C","POOR":"FFC7CE"}
        for i, d in enumerate(report["domain_summary"]):
            r = start_row + 2 + i
            vals = [
                d["domain"],
                d["total_controls"],
                d["covered_controls"],
                d["uncovered_controls"],
                d["coverage_pct"] / 100,
                d.get("full_matches", 0),
                d.get("partial_matches", 0),
                d.get("indirect_matches", 0),
                d["coverage_status"],
                d.get("coverage_note", ""),
            ]
            write_data_row(ws1, r, vals, alt=(i % 2 == 1))
            # Status cell colour
            sc = ws1.cell(row=r, column=9)
            bg = sc_map.get(d["coverage_status"], "FFFFFF")
            sc.fill = fill(bg)
            sc.font = body_font(bold=True)
            # Percentage format
            ws1.cell(row=r, column=5).number_format = "0%"
            # Colour FULL cell green
            fc = ws1.cell(row=r, column=6)
            if d.get("full_matches", 0) > 0:
                fc.fill = fill("C6EFCE")
                fc.font = body_font(bold=True, color="276221")
            # Colour PARTIAL amber
            pc = ws1.cell(row=r, column=7)
            if d.get("partial_matches", 0) > 0:
                pc.fill = fill("FFEB9C")
                pc.font = body_font(bold=True, color="9C5700")
            # Colour INDIRECT grey
            ic = ws1.cell(row=r, column=8)
            if d.get("indirect_matches", 0) > 0:
                ic.fill = fill("EDEDED")
                ic.font = body_font(bold=True, color="595959")

        set_col_widths(ws1, {"A":30,"B":8,"C":10,"D":12,
                              "E":12,"F":8,"G":10,"H":11,"I":12,"J":38})
        ws1.column_dimensions["A"].width = 30
        freeze(ws1, "A2")
        # Extend domain section merge to 10 cols
        ws1.merge_cells(start_row=start_row, start_column=1,
                        end_row=start_row, end_column=10)

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 2 — Control Mappings
        # ═══════════════════════════════════════════════════════════════════════
        ws2 = wb.create_sheet("Control Mappings")
        title_row(ws2, f"Control → Policy Mappings  |  {app}", 14)
        hdrs = ["Control ID","Control Text","Domain","Framework","Subdomain",
                "Risk Level","Is Covered","Policy ID","Policy Name",
                "Policy Category","Impact","Similarity Score","Coverage","Relevance"]
        write_header_row(ws2, 2, hdrs)
        freeze(ws2, "A3")

        row = 3
        for cm in report["control_mappings"]:
            if cm["matches"]:
                for m in cm["matches"]:
                    relevance = ("HIGH"   if m["similarity_score"] >= 0.82 else
                                 "MEDIUM" if m["similarity_score"] >= 0.60 else "LOW")
                    vals = [cm["control_id"], cm["control_text"], cm["domain"],
                            cm["framework"], cm.get("subdomain",""),
                            cm["risk_level"], "Yes" if cm["is_covered"] else "No",
                            m["policy_id"], m["policy_name"],
                            m.get("policy_category",""),
                            m.get("impact",""),
                            m["similarity_score"], m["coverage"], relevance]
                    write_data_row(ws2, row, vals, alt=(row % 2 == 0))
                    # Coverage colour — now column 13 (impact added)
                    cov_cell = ws2.cell(row=row, column=13)
                    if m["coverage"] == "FULL":
                        color_cell(cov_cell, C_FULL, C_FULL_FG)
                    elif m["coverage"] == "PARTIAL":
                        color_cell(cov_cell, C_PARTIAL, C_PARTIAL_FG)
                    else:
                        color_cell(cov_cell, C_INDIRECT, C_INDIRECT_FG)
                    # Risk colour
                    risk_cell = ws2.cell(row=row, column=6)
                    if cm["risk_level"] == "HIGH":
                        color_cell(risk_cell, C_HIGH, C_HIGH_FG)
                    elif cm["risk_level"] == "MEDIUM":
                        color_cell(risk_cell, C_MEDIUM, C_MEDIUM_FG)
                    else:
                        color_cell(risk_cell, C_LOW, C_LOW_FG)
                    # Score number format
                    ws2.cell(row=row, column=11).number_format = "0.0000"
                    row += 1
            else:
                vals = [cm["control_id"], cm["control_text"], cm["domain"],
                        cm["framework"], cm.get("subdomain",""),
                        cm["risk_level"], "No",
                        "","","","","",""]
                write_data_row(ws2, row, vals, alt=(row % 2 == 0))
                # Highlight whole row as uncovered
                for col in range(1, 14):
                    c = ws2.cell(row=row, column=col)
                    c.fill = fill(C_UNCOVERED)
                row += 1

        set_col_widths(ws2, {"A":12,"B":45,"C":20,"D":12,"E":12,"F":11,
                              "G":11,"H":10,"I":40,"J":18,"K":15,"L":14,"M":12,"N":11})
        ws2.auto_filter.ref = f"A2:N{row-1}"

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 3 — Policy Mappings
        # ═══════════════════════════════════════════════════════════════════════
        ws3 = wb.create_sheet("Policy Mappings")
        title_row(ws3, f"Policy → Control Mappings  |  {app}", 7)
        write_header_row(ws3, 2,
            ["Policy ID","Policy Name","Policy Category",
             "Is Orphan","Matched Control ID","Match Count"])
        freeze(ws3, "A3")

        row = 3
        for pm in report["policy_mappings"]:
            if pm["matched_controls"]:
                for ctrl_id in pm["matched_controls"]:
                    vals = [pm["policy_id"], pm["policy_name"],
                            pm.get("policy_category",""),
                            "No", ctrl_id, pm["match_count"]]
                    write_data_row(ws3, row, vals, alt=(row % 2 == 0))
                    row += 1
            else:
                vals = [pm["policy_id"], pm["policy_name"],
                        pm.get("policy_category",""), "Yes", "", 0]
                write_data_row(ws3, row, vals, alt=(row % 2 == 0))
                orphan_cell = ws3.cell(row=row, column=4)
                color_cell(orphan_cell, C_UNCOVERED, C_UNCOVERED_FG)
                row += 1

        set_col_widths(ws3, {"A":12,"B":45,"C":20,"D":11,"E":20,"F":13})
        ws3.auto_filter.ref = f"A2:F{row-1}"

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 4 — Uncovered Controls
        # ═══════════════════════════════════════════════════════════════════════
        ws4 = wb.create_sheet("Uncovered Controls")
        title_row(ws4, f"Uncovered Controls — No Matching OOTB Policy  |  {app}", 8)
        write_header_row(ws4, 2,
            ["Control ID","Control Text","Domain","Framework",
             "Subdomain","Risk Level","Reason"])
        freeze(ws4, "A3")

        row = 3
        for i, uc in enumerate(report["uncovered_controls"]):
            vals = [uc["control_id"], uc["control_text"], uc["domain"],
                    uc["framework"], uc.get("subdomain",""),
                    uc["risk_level"],
                    uc.get("reason","No OOTB policy matches at or above threshold")]
            write_data_row(ws4, row, vals, alt=(i % 2 == 1))
            risk_cell = ws4.cell(row=row, column=6)
            if uc["risk_level"] == "HIGH":
                color_cell(risk_cell, C_HIGH, C_HIGH_FG)
            elif uc["risk_level"] == "MEDIUM":
                color_cell(risk_cell, C_MEDIUM, C_MEDIUM_FG)
            else:
                color_cell(risk_cell, C_LOW, C_LOW_FG)
            row += 1

        if len(report["uncovered_controls"]) == 0:
            ws4.cell(row=3, column=1,
                     value="✅  All controls have at least one matching policy.").font = body_font(bold=True, color="276221")

        set_col_widths(ws4, {"A":12,"B":50,"C":22,"D":12,"E":12,"F":11,"G":45})
        ws4.auto_filter.ref = f"A2:G{max(row-1, 3)}"

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 5 — Orphan Policies
        # ═══════════════════════════════════════════════════════════════════════
        ws5 = wb.create_sheet("Orphan Policies")
        title_row(ws5, f"Orphan Policies — Not Required by Any Standard  |  {app}", 5)
        write_header_row(ws5, 2,
            ["Policy ID","Policy Name","Policy Category","Reason"])
        freeze(ws5, "A3")

        row = 3
        for i, op in enumerate(report["orphan_policies"]):
            vals = [op["policy_id"], op["policy_name"],
                    op.get("policy_category",""),
                    op.get("reason","No org control maps to this OOTB policy")]
            write_data_row(ws5, row, vals, alt=(i % 2 == 1))
            row += 1

        if len(report["orphan_policies"]) == 0:
            ws5.cell(row=3, column=1,
                     value="✅  All policies are referenced by at least one control.").font = body_font(bold=True, color="276221")

        set_col_widths(ws5, {"A":12,"B":45,"C":20,"D":50})
        ws5.auto_filter.ref = f"A2:D{max(row-1, 3)}"

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 6 — Relationships
        # ═══════════════════════════════════════════════════════════════════════
        ws6 = wb.create_sheet("Relationships")
        title_row(ws6, f"Mapping Relationship Types  |  {app}", 3)
        rel = report["relationships"]

        # 1-to-Many: derived directly from control_mappings (same source as One-to-Many sheet)
        # A control is 1-to-many if it has more than 1 match regardless of overlap
        one_to_many_direct = [
            cm for cm in report["control_mappings"] if len(cm["matches"]) > 1
        ]

        # Build sections — 1-to-1 and many-to-* from classify, 1-to-many from mappings
        sections = [
            {
                "label":       "1-to-1",
                "description": "1 control maps to exactly 1 policy",
                "col_hdrs":    ["Control ID", "Policy ID"],
                "count":       len(rel["one_to_one"]),
                "rows": [[item["control"], item["policy"]]
                         for item in rel["one_to_one"]],
            },
            {
                "label":       "1-to-Many",
                "description": "1 control maps to multiple policies (same as One-to-Many sheet)",
                "col_hdrs":    ["Control ID", "Control Text", "Policy IDs", "Policy Count"],
                "count":       len(one_to_many_direct),
                "rows": [
                    [
                        cm["control_id"],
                        cm["control_text"],
                        ", ".join(m["policy_id"] for m in cm["matches"]),
                        len(cm["matches"]),
                    ]
                    for cm in one_to_many_direct
                ],
            },
            {
                "label":       "Many-to-1",
                "description": "Multiple controls map to the same policy",
                "col_hdrs":    ["Policy ID", "Control IDs", "Control Count"],
                "count":       len(rel["many_to_one"]),
                "rows": [[item["policy"],
                          ", ".join(item["controls"]),
                          len(item["controls"])]
                         for item in rel["many_to_one"]],
            },
            {
                "label":       "Many-to-Many",
                "description": "Cluster of controls share a cluster of policies",
                "col_hdrs":    ["Control IDs", "Policy IDs", "Size"],
                "count":       len(rel["many_to_many"]),
                "rows": [[", ".join(item["controls"]),
                          ", ".join(item["policies"]),
                          f"{len(item['controls'])} x {len(item['policies'])}"]
                         for item in rel["many_to_many"]],
            },
        ]

        current_row = 3
        for sec in sections:
            # Section header
            ncols = len(sec["col_hdrs"])
            ws6.merge_cells(start_row=current_row, start_column=1,
                            end_row=current_row, end_column=max(ncols, 3))
            c = ws6.cell(row=current_row, column=1,
                         value=f"{sec['label']}  ({sec['count']})  —  {sec['description']}")
            c.font      = hdr_font(size=11, color="FFFFFF")
            c.fill      = fill(C_HEADER_MID)
            c.alignment = Alignment(horizontal="left", vertical="center")
            ws6.row_dimensions[current_row].height = 22
            current_row += 1

            write_header_row(ws6, current_row, sec["col_hdrs"])
            current_row += 1

            if not sec["rows"]:
                ws6.cell(row=current_row, column=1,
                         value="No relationships of this type.").font = body_font(color="595959")
                current_row += 1
            else:
                for i, row_vals in enumerate(sec["rows"]):
                    write_data_row(ws6, current_row, row_vals, alt=(i % 2 == 1))
                    # Wrap long policy ID lists
                    for col in range(1, len(row_vals) + 1):
                        ws6.cell(row=current_row, column=col).alignment = wrap_align()
                    current_row += 1

            current_row += 1  # blank spacer

        set_col_widths(ws6, {"A":20,"B":50,"C":50,"D":12})

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 7 — One-to-Many (controls that map to multiple policies)
        # ═══════════════════════════════════════════════════════════════════════
        ws7 = wb.create_sheet("One-to-Many")
        title_row(ws7, f"One Control → Multiple Policies  |  {app}", 6)
        write_header_row(ws7, 2,
            ["Control ID", "Control Text", "Domain", "Framework",
             "Policy ID", "Policy Name", "Policy Category", "Impact",
             "Similarity Score", "Coverage", "Match Rank"])
        freeze(ws7, "A3")

        row = 3
        # Only controls that matched MORE than one policy
        multi_match = [cm for cm in report["control_mappings"] if len(cm["matches"]) > 1]
        for cm in multi_match:
            for rank, m in enumerate(cm["matches"], 1):
                relevance = ("HIGH"   if m["similarity_score"] >= 0.82 else
                             "MEDIUM" if m["similarity_score"] >= 0.60 else "LOW")
                vals = [
                    cm["control_id"],
                    cm["control_text"],
                    cm["domain"],
                    cm["framework"],
                    m["policy_id"],
                    m["policy_name"],
                    m.get("policy_category", ""),
                    m.get("impact", ""),
                    m["similarity_score"],
                    m["coverage"],
                    f"#{rank}",
                ]
                write_data_row(ws7, row, vals, alt=(row % 2 == 0))
                # Coverage colour — now column 10 (impact added)
                cov_cell = ws7.cell(row=row, column=10)
                if m["coverage"] == "FULL":
                    color_cell(cov_cell, C_FULL, C_FULL_FG)
                elif m["coverage"] == "PARTIAL":
                    color_cell(cov_cell, C_PARTIAL, C_PARTIAL_FG)
                else:
                    color_cell(cov_cell, C_INDIRECT, C_INDIRECT_FG)
                ws7.cell(row=row, column=8).number_format = "0.0000"
                row += 1
            # Blank separator row between controls
            row += 1

        if not multi_match:
            ws7.cell(row=3, column=1,
                     value="No controls matched more than one policy.").font = body_font(color="595959")

        set_col_widths(ws7, {"A":12,"B":45,"C":22,"D":12,
                              "E":12,"F":40,"G":18,"H":15,"I":14,"J":12,"K":10})
        ws7.auto_filter.ref = f"A2:K{max(row-1, 3)}"

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 8 — Mind Map (text-based tree layout in Excel)
        # ═══════════════════════════════════════════════════════════════════════
        ws8 = wb.create_sheet("Mind Map")
        title_row(ws8, f"Control → Policy Mind Map  |  {app}", 8)

        # Colour palette for mind map levels
        C_DOMAIN    = "1F3864"   # navy   — domain root
        C_DOMAIN_FG = "FFFFFF"
        C_CONTROL   = "2E75B6"   # blue   — control node
        C_CTRL_FG   = "FFFFFF"
        C_FULL_N    = "375623"   # dark green — FULL match node
        C_FULL_NFG  = "FFFFFF"
        C_PART_N    = "7F6000"   # dark amber — PARTIAL match node
        C_PART_NFG  = "FFFFFF"
        C_IND_N     = "595959"   # grey — INDIRECT node
        C_IND_NFG   = "FFFFFF"
        C_UNMAP     = "843C0C"   # red-brown — unmapped
        C_UNMAP_FG  = "FFFFFF"

        # Column layout:
        # A        = Domain
        # B        = Control ID
        # C        = Control Text (truncated)
        # D        = connector arrow
        # E        = Policy ID
        # F        = Policy Name
        # G        = Coverage badge

        # Header row — 8 columns now (added Impact)
        # Col: A=indent  B=ControlID  C=ControlText  D=arrow
        #      E=PolicyID  F=PolicyName  G=Coverage  H=Impact
        mm_hdrs = ["", "Control ID", "Control Text",
                   "", "Policy ID", "Policy Name", "Coverage", "Impact"]
        write_header_row(ws8, 2, mm_hdrs, bg=C_HEADER_DARK, fg="FFFFFF")
        ws8.row_dimensions[2].height = 20

        # Group controls by domain
        from collections import defaultdict
        domain_map = defaultdict(list)
        for cm in report["control_mappings"]:
            domain_map[cm["domain"]].append(cm)

        mm_row = 3
        for domain, controls in sorted(domain_map.items()):
            # Domain header row — spans all columns
            ws8.merge_cells(start_row=mm_row, start_column=1,
                            end_row=mm_row, end_column=8)
            dc = ws8.cell(row=mm_row, column=1,
                          value=f"▶  {domain.upper()}")
            dc.font      = Font(name="Arial", size=11, bold=True, color=C_DOMAIN_FG)
            dc.fill      = fill(C_DOMAIN)
            dc.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
            ws8.row_dimensions[mm_row].height = 22
            mm_row += 1

            for cm in controls:
                n_matches = len(cm["matches"])

                if n_matches == 0:
                    # Unmapped control — single row
                    ws8.cell(row=mm_row, column=1, value="")
                    ctrl_id_cell = ws8.cell(row=mm_row, column=2,
                                            value=cm["control_id"])
                    ctrl_id_cell.font = Font(name="Arial", size=10,
                                             bold=True, color=C_UNMAP_FG)
                    ctrl_id_cell.fill = fill(C_UNMAP)
                    ctrl_id_cell.alignment = wrap_align("center")

                    ctrl_txt = ws8.cell(row=mm_row, column=3,
                                        value=cm["control_text"])
                    ctrl_txt.font = body_font(color=C_UNMAP)
                    ctrl_txt.alignment = wrap_align()
                    ws8.row_dimensions[mm_row].height = None  # auto height

                    ws8.cell(row=mm_row, column=4, value="✘ NO MATCH").font =                         Font(name="Arial", size=10, bold=True, color=C_UNMAP)
                    mm_row += 1

                else:
                    # One row per matched policy
                    for idx, m in enumerate(cm["matches"]):
                        # Control ID + text only on first match row
                        if idx == 0:
                            ctrl_id_cell = ws8.cell(row=mm_row, column=2,
                                                    value=cm["control_id"])
                            ctrl_id_cell.font      = Font(name="Arial", size=10,
                                                          bold=True, color=C_CTRL_FG)
                            ctrl_id_cell.fill      = fill(C_CONTROL)
                            ctrl_id_cell.alignment = wrap_align("center")

                            ctrl_txt = ws8.cell(row=mm_row, column=3,
                                                value=cm["control_text"])
                            ctrl_txt.font      = body_font(bold=True)
                            ctrl_txt.alignment = wrap_align()

                            # Merge control text cell vertically if multiple matches
                            if n_matches > 1:
                                try:
                                    ws8.merge_cells(
                                        start_row=mm_row, start_column=2,
                                        end_row=mm_row + n_matches - 1,
                                        end_column=2)
                                    ws8.merge_cells(
                                        start_row=mm_row, start_column=3,
                                        end_row=mm_row + n_matches - 1,
                                        end_column=3)
                                except Exception:
                                    pass  # skip merge errors

                        # Connector arrow
                        arrow = ws8.cell(row=mm_row, column=4,
                                         value="──►" if idx == 0 else "   ►")
                        arrow.font      = Font(name="Arial", size=10, color="2E75B6")
                        arrow.alignment = Alignment(horizontal="center",
                                                    vertical="center")

                        # Policy ID
                        pol_id_cell = ws8.cell(row=mm_row, column=5,
                                               value=m["policy_id"])
                        pol_id_cell.font      = body_font(bold=True)
                        pol_id_cell.alignment = wrap_align("center")

                        # Policy name
                        pol_name_cell = ws8.cell(row=mm_row, column=6,
                                                 value=m["policy_name"])
                        pol_name_cell.font      = body_font()
                        pol_name_cell.alignment = wrap_align()

                        # Coverage badge
                        cov_val = m["coverage"]
                        cov_cell = ws8.cell(row=mm_row, column=7, value=cov_val)
                        if cov_val == "FULL":
                            color_cell(cov_cell, C_FULL_N, C_FULL_NFG)
                        elif cov_val == "PARTIAL":
                            color_cell(cov_cell, C_PART_N, C_PART_NFG)
                        else:
                            color_cell(cov_cell, C_IND_N, C_IND_NFG)
                        cov_cell.alignment = Alignment(horizontal="center",
                                                       vertical="center")

                        # Impact cell — col 8
                        impact_val = m.get("impact", "")
                        imp_cell = ws8.cell(row=mm_row, column=8, value=impact_val)
                        imp_cell.font      = body_font(bold=bool(impact_val))
                        imp_cell.alignment = Alignment(horizontal="center",
                                                       vertical="center")
                        # Colour impact by severity if value present
                        if impact_val:
                            iv = impact_val.upper()
                            if any(x in iv for x in ("HIGH","CRITICAL","SEVERE")):
                                imp_cell.fill = fill(C_HIGH)
                                imp_cell.font = Font(name="Arial", size=10,
                                                     bold=True, color=C_HIGH_FG)
                            elif any(x in iv for x in ("MEDIUM","MODERATE")):
                                imp_cell.fill = fill(C_MEDIUM)
                                imp_cell.font = Font(name="Arial", size=10,
                                                     bold=True, color=C_MEDIUM_FG)
                            elif any(x in iv for x in ("LOW","MINOR","INFO")):
                                imp_cell.fill = fill(C_LOW)
                                imp_cell.font = Font(name="Arial", size=10,
                                                     bold=True, color=C_LOW_FG)

                        # Alt row shading for non-coloured cells
                        for col in [1, 4, 5, 6]:
                            c = ws8.cell(row=mm_row, column=col)
                            if not c.value:
                                c.fill = fill(C_ALT_ROW if mm_row % 2 == 0
                                              else "FFFFFF")
                            c.border = border

                        ws8.row_dimensions[mm_row].height = 18
                        mm_row += 1

            # Spacer between domains
            mm_row += 1

        set_col_widths(ws8, {"A":3, "B":13, "C":42,
                              "D":6, "E":13, "F":42, "G":11, "H":14})
        ws8.freeze_panes = "A3"

        # ═══════════════════════════════════════════════════════════════════════
        # SHEET 9 — Domain Graph View
        # Visual tree: Security Domain → Policy Domain → Policies
        # ═══════════════════════════════════════════════════════════════════════
        ws9 = wb.create_sheet("Domain Graph")
        title_row(ws9, f"Domain Mapping Graph  |  {app}", 5)

        # Colour palette for tree levels
        C_L1 = "1F3864"   # Level 1 — Security Domain (navy)
        C_L2 = "2E75B6"   # Level 2 — Policy Domain   (blue)
        C_L3 = "D6E4F0"   # Level 3 — Policy name     (light blue)
        C_UNK = "7F6000"  # Unknown domain             (amber)

        # Column layout:
        # A = Security Domain (std domain)
        # B = connector  →
        # C = Policy Domain (policy category)
        # D = connector  →
        # E = Policy ID
        # F = Policy Name
        # G = Coverage (from best match)
        # H = Match Count

        ws9.column_dimensions["A"].width = 28
        ws9.column_dimensions["B"].width = 4
        ws9.column_dimensions["C"].width = 22
        ws9.column_dimensions["D"].width = 4
        ws9.column_dimensions["E"].width = 14
        ws9.column_dimensions["F"].width = 45
        ws9.column_dimensions["G"].width = 12
        ws9.column_dimensions["H"].width = 10

        # Header row
        g_hdrs = ["Security Domain","","Policy Domain","","Policy ID",
                  "Policy Name","Best Coverage","Match Count"]
        write_header_row(ws9, 2, g_hdrs, bg=C_HEADER_DARK, fg="FFFFFF")
        ws9.freeze_panes = "A3"

        # Build tree structure from control_mappings
        # tree: { std_domain → { pol_category → [ {policy_id, policy_name, coverage, match_count} ] } }
        from collections import defaultdict
        tree = defaultdict(lambda: defaultdict(list))
        pol_best = {}  # policy_id → best coverage across all controls

        for cm in report["control_mappings"]:
            std_dom = cm["domain"]
            for m in cm["matches"]:
                pol_cat = m.get("policy_category", "Uncategorised") or "Uncategorised"
                pid     = m["policy_id"]
                pname   = m["policy_name"]
                cov     = m["coverage"]

                # Track best coverage per policy
                cov_rank = {"FULL": 3, "PARTIAL": 2, "INDIRECT": 1}
                prev = pol_best.get(pid, {})
                if cov_rank.get(cov, 0) >= cov_rank.get(prev.get("coverage",""), 0):
                    pol_best[pid] = {
                        "policy_id": pid,
                        "policy_name": pname,
                        "policy_category": pol_cat,
                        "coverage": cov,
                        "std_domains": set(),
                    }
                pol_best[pid]["std_domains"].add(std_dom)
                tree[std_dom][pol_cat].append(pid)

        # De-duplicate policy IDs per branch
        for std_dom in tree:
            for pol_cat in tree[std_dom]:
                tree[std_dom][pol_cat] = sorted(set(tree[std_dom][pol_cat]))

        # Also show domain pairs that have NO matches (empty branches)
        for std_dom_key, allowed_cats in DOMAIN_PAIRS.items():
            # Find matching control domain
            matching = [cm["domain"] for cm in report["control_mappings"]
                       if std_dom_key in cm["domain"].lower()
                       or cm["domain"].lower() in std_dom_key]
            display_dom = matching[0] if matching else std_dom_key.title()
            for cat in allowed_cats:
                if cat.title() not in tree.get(display_dom, {}):
                    if display_dom not in tree:
                        tree[display_dom] = {}
                    if cat.title() not in tree[display_dom]:
                        tree[display_dom][cat.title()] = []  # empty branch

        g_row = 3
        cov_colors = {"FULL": C_FULL, "PARTIAL": C_PARTIAL,
                      "INDIRECT": C_INDIRECT_FG, "": "AAAAAA"}
        cov_fg     = {"FULL": C_FULL_FG, "PARTIAL": C_PARTIAL_FG,
                      "INDIRECT": "FFFFFF", "": "FFFFFF"}

        for std_dom in sorted(tree.keys()):
            pol_cats = tree[std_dom]
            total_policies = sum(len(v) for v in pol_cats.values())

            # ── Level 1: Security Domain header ──────────────────────────────
            n_cat_rows = max(total_policies, 1)
            ws9.merge_cells(start_row=g_row, start_column=1,
                            end_row=g_row + n_cat_rows - 1, end_column=1)
            l1 = ws9.cell(row=g_row, column=1, value=f"▶  {std_dom}")
            l1.font      = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            l1.fill      = fill(C_L1)
            l1.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1, wrap_text=True)

            cat_start_row = g_row

            for pol_cat in sorted(pol_cats.keys()):
                pol_ids   = pol_cats[pol_cat]
                n_pol     = max(len(pol_ids), 1)
                is_unknown = pol_cat.lower() not in {
                    d.lower()
                    for ds in DOMAIN_PAIRS.values() for d in ds
                }

                # ── Level 2: Policy Domain ────────────────────────────────────
                ws9.merge_cells(start_row=cat_start_row, start_column=3,
                                end_row=cat_start_row + n_pol - 1, end_column=3)
                # Arrow from L1 to L2
                ws9.merge_cells(start_row=cat_start_row, start_column=2,
                                end_row=cat_start_row + n_pol - 1, end_column=2)
                arr1 = ws9.cell(row=cat_start_row, column=2, value="──►")
                arr1.font      = Font(name="Arial", size=10, color="2E75B6")
                arr1.alignment = Alignment(horizontal="center", vertical="center")

                l2_bg = C_UNK if is_unknown else C_L2
                l2 = ws9.cell(row=cat_start_row, column=3,
                              value=("⚠ " if is_unknown else "") + pol_cat)
                l2.font      = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                l2.fill      = fill(l2_bg)
                l2.alignment = Alignment(horizontal="left", vertical="center",
                                         indent=1, wrap_text=True)

                if not pol_ids:
                    # Empty branch — no policies matched yet
                    arr2e = ws9.cell(row=cat_start_row, column=4, value="──►")
                    arr2e.font = Font(name="Arial", size=10, color="AAAAAA")
                    arr2e.alignment = Alignment(horizontal="center", vertical="center")
                    empty = ws9.cell(row=cat_start_row, column=5,
                                     value="(no matches yet)")
                    empty.font = Font(name="Arial", size=9, color="AAAAAA",
                                      italic=True)
                    for col in [6, 7, 8]:
                        ws9.cell(row=cat_start_row, column=col).fill = fill("F5F5F5")
                    ws9.row_dimensions[cat_start_row].height = 18
                    cat_start_row += 1
                else:
                    # ── Level 3: Individual Policies ─────────────────────────
                    for k, pid in enumerate(pol_ids):
                        pd   = pol_best.get(pid, {})
                        pname = pd.get("policy_name", pid)
                        cov   = pd.get("coverage", "")

                        # Arrow from L2 to L3
                        arr2 = ws9.cell(row=cat_start_row, column=4,
                                        value="──►" if k == 0 else "   ►")
                        arr2.font = Font(name="Arial", size=10, color="2E75B6")
                        arr2.alignment = Alignment(horizontal="center",
                                                   vertical="center")

                        # Policy ID
                        pid_cell = ws9.cell(row=cat_start_row, column=5, value=pid)
                        pid_cell.font      = body_font(bold=True)
                        pid_cell.alignment = wrap_align("center")

                        # Policy Name
                        pn_cell = ws9.cell(row=cat_start_row, column=6, value=pname)
                        pn_cell.font      = body_font()
                        pn_cell.alignment = wrap_align()

                        # Coverage badge
                        cv_bg = cov_colors.get(cov, "AAAAAA")
                        cv_fg = cov_fg.get(cov, "FFFFFF")
                        cov_cell = ws9.cell(row=cat_start_row, column=7, value=cov or "—")
                        cov_cell.fill      = fill(cv_bg)
                        cov_cell.font      = Font(name="Arial", size=10,
                                                  bold=True, color=cv_fg)
                        cov_cell.alignment = Alignment(horizontal="center",
                                                       vertical="center")

                        # Match count (how many controls map to this policy)
                        n_ctrl = len(pd.get("std_domains", set()))
                        mc_cell = ws9.cell(row=cat_start_row, column=8,
                                           value=n_ctrl if n_ctrl else "")
                        mc_cell.font      = body_font()
                        mc_cell.alignment = Alignment(horizontal="center")

                        # Alternating row shading
                        for col in [1, 2, 3, 4]:
                            c = ws9.cell(row=cat_start_row, column=col)
                            if not c.value:
                                c.fill = fill(C_ALT_ROW if cat_start_row % 2 == 0
                                              else "FFFFFF")
                            c.border = border

                        ws9.row_dimensions[cat_start_row].height = 18
                        cat_start_row += 1

            # Blank row between security domains
            cat_start_row += 1
            g_row = cat_start_row

        # Add legend
        legend_row = g_row + 1
        ws9.merge_cells(start_row=legend_row, start_column=1,
                        end_row=legend_row, end_column=8)
        legend = ws9.cell(row=legend_row, column=1,
                          value="Legend:  ▶ Security Domain  ──►  Policy Domain  ──►  Policy  |  "
                                "Green=FULL  Amber=PARTIAL  Grey=INDIRECT  "
                                "Amber header=NEW domain (add to domain_pairs.json)")
        legend.font      = Font(name="Arial", size=9, italic=True, color="595959")
        legend.alignment = Alignment(horizontal="left")

        # ── Save ──────────────────────────────────────────────────────────────
        wb.save(path)
        print(f"📊  Excel report saved to {path}")
        print(f"     Sheets: Summary | Control Mappings | Policy Mappings |")
        print(f"             Uncovered Controls | Orphan Policies | Relationships |")
        print(f"             One-to-Many | Mind Map | Domain Graph")


    def save_csv(self, report: dict, prefix: str = "mapping"):
        """
        Export mapping results to 4 CSV files:

          {prefix}_control_mappings.csv   — every control with all its matched policies
          {prefix}_policy_mappings.csv    — every policy with all its matched controls
          {prefix}_uncovered_controls.csv — controls that have NO matching policy
          {prefix}_orphan_policies.csv    — policies that have NO matching control

        All files include the security domain from the input controls.
        """
        import csv
        app = report.get("app", "")

        # ── 1. Control Mappings ───────────────────────────────────────────────
        # One row per (control, matched_policy) pair.
        # If a control has 3 matches → 3 rows, all with same control fields.
        path1 = f"{prefix}_control_mappings.csv"
        with open(path1, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "app",
                "control_id",
                "control_text",
                "domain",
                "framework",
                "subdomain",
                "risk_level",
                "is_covered",
                "policy_id",
                "policy_name",
                "policy_category",
                "similarity_score",
                "coverage",
                "relevance",
            ])
            for cm in report["control_mappings"]:
                if cm["matches"]:
                    for m in cm["matches"]:
                        relevance = (
                            "HIGH"   if m["similarity_score"] >= 0.82 else
                            "MEDIUM" if m["similarity_score"] >= 0.60 else
                            "LOW"
                        )
                        writer.writerow([
                            app,
                            cm["control_id"],
                            cm["control_text"],
                            cm["domain"],
                            cm["framework"],
                            cm.get("subdomain", ""),
                            cm["risk_level"],
                            cm["is_covered"],
                            m["policy_id"],
                            m["policy_name"],
                            m.get("policy_category", ""),
                            m["similarity_score"],
                            m["coverage"],
                            relevance,
                        ])
                else:
                    # Control has no matches — still include as a row with blanks
                    writer.writerow([
                        app,
                        cm["control_id"],
                        cm["control_text"],
                        cm["domain"],
                        cm["framework"],
                        cm.get("subdomain", ""),
                        cm["risk_level"],
                        False,
                        "", "", "", "", "", "",
                    ])
        print(f"📄  Saved: {path1}")

        # ── 2. Policy Mappings ────────────────────────────────────────────────
        # One row per (policy, matched_control) pair.
        path2 = f"{prefix}_policy_mappings.csv"
        with open(path2, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "app",
                "policy_id",
                "policy_name",
                "policy_category",
                "is_orphan",
                "matched_control_id",
                "match_count",
            ])
            for pm in report["policy_mappings"]:
                if pm["matched_controls"]:
                    for ctrl_id in pm["matched_controls"]:
                        writer.writerow([
                            app,
                            pm["policy_id"],
                            pm["policy_name"],
                            pm.get("policy_category", ""),
                            pm["is_orphan"],
                            ctrl_id,
                            pm["match_count"],
                        ])
                else:
                    writer.writerow([
                        app,
                        pm["policy_id"],
                        pm["policy_name"],
                        pm.get("policy_category", ""),
                        True,
                        "",
                        0,
                    ])
        print(f"📄  Saved: {path2}")

        # ── 3. Uncovered Controls ─────────────────────────────────────────────
        # Standards that have NO matching OOTB policy
        path3 = f"{prefix}_uncovered_controls.csv"
        with open(path3, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "app",
                "control_id",
                "control_text",
                "domain",
                "framework",
                "subdomain",
                "risk_level",
                "reason",
            ])
            for uc in report["uncovered_controls"]:
                writer.writerow([
                    app,
                    uc["control_id"],
                    uc["control_text"],
                    uc["domain"],
                    uc["framework"],
                    uc.get("subdomain", ""),
                    uc["risk_level"],
                    uc.get("reason", "No OOTB policy matches at or above threshold"),
                ])
        count_uc = len(report["uncovered_controls"])
        print(f"📄  Saved: {path3}  ({count_uc} uncovered controls)")

        # ── 4. Orphan Policies ────────────────────────────────────────────────
        # Policies that NO control maps to
        path4 = f"{prefix}_orphan_policies.csv"
        with open(path4, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "app",
                "policy_id",
                "policy_name",
                "policy_category",
                "reason",
            ])
            for op in report["orphan_policies"]:
                writer.writerow([
                    app,
                    op["policy_id"],
                    op["policy_name"],
                    op.get("policy_category", ""),
                    op.get("reason", "No org control maps to this OOTB policy"),
                ])
        count_op = len(report["orphan_policies"])
        print(f"📄  Saved: {path4}  ({count_op} orphan policies)")

        # ── 5. Domain Summary ─────────────────────────────────────────────────
        path5 = f"{prefix}_domain_summary.csv"
        with open(path5, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "app",
                "domain",
                "total_controls",
                "covered_controls",
                "uncovered_controls",
                "coverage_pct",
                "full_matches",
                "partial_matches",
                "indirect_matches",
                "coverage_status",
                "coverage_note",
                "high_risk_controls",
                "uncovered_control_ids",
            ])
            for d in report["domain_summary"]:
                writer.writerow([
                    app,
                    d["domain"],
                    d["total_controls"],
                    d["covered_controls"],
                    d["uncovered_controls"],
                    d["coverage_pct"],
                    d.get("full_matches", 0),
                    d.get("partial_matches", 0),
                    d.get("indirect_matches", 0),
                    d["coverage_status"],
                    d.get("coverage_note", ""),
                    d["high_risk_controls"],
                    ", ".join(d.get("uncovered_control_ids", [])),
                ])
        print(f"📄  Saved: {path5}")
        print(f"\n✅  All CSV outputs saved with prefix: '{prefix}'")


# ── Sample Data Generator ─────────────────────────────────────────────────────

def generate_sample_data():
    """Creates sample controls and policies files for testing."""

    # ── Controls (org standards input) ───────────────────────────────────────
    controls = [
        # Access Control domain
        ["CTR-001","All users must authenticate using MFA before accessing any SaaS application","Access Control","ISO 27001","A.9.4.2",""],
        ["CTR-002","Privileged accounts must use hardware MFA tokens","Access Control","ISO 27001","A.9.4.2","Admin and service accounts"],
        ["CTR-003","Legacy authentication protocols must be disabled organisation-wide","Access Control","ISO 27001","A.9.4.3",""],
        ["CTR-004","Access must be restricted based on user role and least privilege principle","Access Control","ISO 27001","A.9.2.3",""],
        ["CTR-005","All third-party application integrations must be reviewed and authorised","Access Control","ISO 27001","A.15.2.1",""],
        # Data Protection domain
        ["CTR-006","Sensitive PII and financial data must not be shared with external parties without approval","Data Protection","ISO 27001","A.8.2.3",""],
        ["CTR-007","All sensitive data must be encrypted at rest using AES-256 or equivalent","Data Protection","ISO 27001","A.10.1.1",""],
        ["CTR-008","Data Loss Prevention controls must be configured to detect sensitive data in transit","Data Protection","NIST CSF","PR.DS-5",""],
        # Logging & Monitoring domain
        ["CTR-009","All user authentication events must be logged and retained for minimum 12 months","Logging & Monitoring","ISO 27001","A.12.4.1",""],
        ["CTR-010","All privileged actions and admin configuration changes must be audited","Logging & Monitoring","ISO 27001","A.12.4.3",""],
        ["CTR-011","Security events must be forwarded to SIEM within 5 minutes","Logging & Monitoring","NIST CSF","DE.CM-1",""],
        # Identity Management domain
        ["CTR-012","User provisioning and deprovisioning must be automated via SCIM or equivalent","Identity Management","ISO 27001","A.9.2.1",""],
        ["CTR-013","Passwords must meet complexity requirements: min 12 chars, uppercase, numbers, symbols","Identity Management","ISO 27001","A.9.4.3",""],
        ["CTR-014","Session idle timeout must not exceed 15 minutes for privileged sessions","Identity Management","ISO 27001","A.9.4.2",""],
        # Incident Response domain
        ["CTR-015","Anomalous login behaviour must trigger automated alerts within 10 minutes","Incident Response","NIST CSF","DE.AE-2",""],
        ["CTR-016","Suspected account compromise must initiate automatic session termination","Incident Response","NIST CSF","RS.RP-1",""],
        # Compliance domain
        ["CTR-017","All SaaS platforms must provide audit reports for compliance reviews","Compliance","ISO 27001","A.18.2.3",""],
        ["CTR-018","Data residency requirements must be enforced — EU data must not leave EU region","Compliance","GDPR","Art.46",""],
    ]

    with open("sample_controls.csv","w",newline="",encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["control_id","control_text","domain","framework","subdomain","description"])
        w.writerows(controls)

    # ── Policies (OOTB for Microsoft 365) — NO framework references ──────────
    policies = [
        ["POL-001","Enable Multi-Factor Authentication for all users","Identity","Require MFA at every login for all Microsoft 365 accounts"],
        ["POL-002","Block legacy authentication protocols","Identity","Disable Basic Auth SMTP Auth IMAP POP3 across all services"],
        ["POL-003","Configure Conditional Access policies","Access Control","Enforce access policies based on user location device compliance and sign-in risk"],
        ["POL-004","Enable audit log search and retention","Logging","Enable unified audit log and retain for 90 days minimum"],
        ["POL-005","Disable anonymous sharing in SharePoint and OneDrive","Data Protection","Prevent unauthenticated external access to files and folders"],
        ["POL-006","Enable Safe Links and Safe Attachments","Threat Protection","Scan links and attachments for malware and phishing in real time"],
        ["POL-007","Enforce password complexity and expiration policy","Identity","Set minimum password length complexity and rotation requirements"],
        ["POL-008","Restrict mailbox auto-forwarding to external domains","Data Protection","Block rules that automatically forward emails outside the organisation"],
        ["POL-009","Enable Microsoft 365 DLP policies","Data Protection","Detect and block transmission of sensitive data types in email and files"],
        ["POL-010","Enable Microsoft Secure Score recommendations","Governance","Track and implement Microsoft security baseline recommendations"],
        ["POL-011","Enable Azure AD SCIM provisioning","Identity","Automate user lifecycle management via SCIM protocol"],
        ["POL-012","Configure session timeout for inactive sessions","Session","Automatically sign out users after period of inactivity"],
        ["POL-013","Enable Microsoft Sentinel log streaming","Logging","Forward security events from Microsoft 365 to SIEM in real time"],
        ["POL-014","Enable Identity Protection risk-based sign-in policies","Threat Detection","Automatically challenge or block risky sign-in attempts"],
        ["POL-015","Enable Microsoft Purview eDiscovery and retention","Compliance","Configure data retention labels and legal hold capabilities"],
        ["POL-016","Configure data residency settings","Compliance","Set tenant data location to specific geographic region"],
        ["POL-017","Enable Hardware FIDO2 security key authentication","Identity","Allow and enforce phishing-resistant hardware security keys for privileged users"],
        ["POL-018","Enable Privileged Identity Management (PIM)","Access Control","Just-in-time privileged access with approval workflows and time limits"],
        ["POL-019","Configure Microsoft Defender for Cloud Apps anomaly detection","Threat Detection","Alert on impossible travel unusual file downloads and suspicious admin actions"],
        ["POL-020","Restrict third-party OAuth app permissions","Access Control","Require admin consent for third-party apps requesting Microsoft 365 permissions"],
    ]

    # Add impact column to sample policies
    policies_with_impact = []
    impact_map = {
        "Identity":         "HIGH",
        "Access Control":   "HIGH",
        "Data Protection":  "HIGH",
        "Logging":          "MEDIUM",
        "Compliance":       "MEDIUM",
        "Session":          "MEDIUM",
        "Threat Detection": "HIGH",
        "Threat Protection":"HIGH",
        "Governance":       "LOW",
    }
    for p in policies:
        cat = p[2]
        impact = impact_map.get(cat, "MEDIUM")
        policies_with_impact.append(p + [impact])

    with open("sample_policies_m365.csv","w",newline="",encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["policy_id","policy_name","category","description","impact"])
        w.writerows(policies_with_impact)

    print("📄  Generated sample_controls.csv      (18 controls, 6 domains)")
    print("📄  Generated sample_policies_m365.csv (20 policies, Microsoft 365)\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def parse_args():
    """
    Parse command line arguments.

    Usage:
      python sspm_mapper.py \
        --controls  your_controls.csv \
        --policies  your_policies.csv \
        --app       "Microsoft 365"   \
        --out       my_report         \
        --verbose

    All arguments are optional — defaults to sample data.
    """
    import argparse
    p = argparse.ArgumentParser(
        description="SSPM Policy Mapper — maps org security controls to OOTB policies",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    p.add_argument("--controls", "-c",
                   default=None,
                   help="Path to controls CSV/JSON file\n"
                        "Columns: control_id, control_text, domain, framework, subdomain")
    p.add_argument("--policies", "-p",
                   default=None,
                   help="Path to OOTB policies CSV/JSON file\n"
                        "Columns: policy_id, policy_name, category, description")
    p.add_argument("--app", "-a",
                   default=None,
                   help="SaaS application name (used as report title and output prefix)\n"
                        "Example: --app \"Microsoft 365\"")
    p.add_argument("--out", "-o",
                   default=None,
                   help="Output file prefix (default: derived from app name)\n"
                        "Example: --out salesforce  →  salesforce_report.xlsx")
    p.add_argument("--threshold", "-t",
                   type=float, default=0.40,
                   help="Minimum similarity score to include a match (default: 0.40)")
    p.add_argument("--topk", "-k",
                   type=int, default=5,
                   help="Max number of policy matches per control (default: 5)")
    p.add_argument("--verbose", "-v",
                   action="store_true",
                   help="Show all control→policy match details in console output")
    p.add_argument("--model", "-m",
                   default=None,
                   help="Path to local SecBERT model folder\n"
                        "Example: --model ./secbert_clean")
    p.add_argument("--clear-cache",
                   action="store_true",
                   help="Delete all cached policy embeddings before running\n"
                        "Use when switching policy files or SecBERT model")
    p.add_argument("--no-domain-filter",
                   action="store_true",
                   help="Disable domain pairing filter — compare all controls vs all policies\n"
                        "Use this to see all possible matches regardless of domain")
    p.add_argument("--domain-pairs", "-d",
                   default=None,
                   help="Path to JSON file with custom domain pairing map\n"
                        "Format: {\"Standard Domain\": [\"Policy Domain 1\", ...]}")
    return p.parse_args()


def main():
    import sys
    args = parse_args()

    print("="*70)
    print("  🛡️  SSPM Policy Mapper v4 — Bidirectional Gap Analysis")
    print("="*70+"\n")

    # ── Resolve file paths ────────────────────────────────────────────────────
    # If no files passed → generate and use sample data
    controls_file = args.controls
    policies_file = args.policies

    if not controls_file or not policies_file:
        if not Path("sample_controls.csv").exists():
            generate_sample_data()
        controls_file = controls_file or "sample_controls.csv"
        policies_file = policies_file or "sample_policies_m365.csv"
        print(f"ℹ️   No files specified — using sample data.")
        print(f"     Controls : {controls_file}")
        print(f"     Policies : {policies_file}\n")
        print(f"     To use your own files:")
        print(f"     python sspm_mapper.py --controls your_controls.csv \\")
        print(f'                           --policies your_policies.csv \\')
        print(f'                           --app "Your App Name"\n')

    # ── Derive app name ───────────────────────────────────────────────────────
    # Priority: --app argument > filename stem > "Unknown App"
    if args.app:
        app_name = args.app
    else:
        # Derive from policy filename: "sample_policies_m365.csv" → "M365"
        stem = Path(policies_file).stem               # e.g. "sample_policies_m365"
        stem = stem.replace("sample_policies_", "")   # → "m365"
        stem = stem.replace("policies_", "")          # → "m365"
        stem = stem.replace("_policies", "")
        stem = stem.replace("_", " ").strip()
        app_name = stem.title() if stem else "Unknown App"
        print(f"ℹ️   No --app specified — using '{app_name}' derived from filename.")
        print(f"     Pass --app \"Your App Name\" for a proper title.\n")

    # ── Derive output prefix ──────────────────────────────────────────────────
    if args.out:
        out_prefix = args.out
    else:
        out_prefix = app_name.lower().replace(" ", "_").replace("/", "_")

    print(f"  App     : {app_name}")
    print(f"  Controls: {controls_file}")
    print(f"  Policies: {policies_file}")
    print(f"  Output  : {out_prefix}_report.xlsx\n")

    # ── Run mapper ────────────────────────────────────────────────────────────
    mapper = SSPMMapper(model_path=args.model)

    # Clear cache if requested
    if getattr(args, 'clear_cache', False):
        mapper.cache.clear()

    # Handle domain filter options
    if getattr(args, 'no_domain_filter', False):
        import sspm_mapper as _sm
        _sm.DOMAIN_PAIRS.clear()
        print("ℹ️   Domain filter disabled — all controls vs all policies.\n")
    elif getattr(args, 'domain_pairs', None):
        import json as _json, sspm_mapper as _sm
        custom = _json.load(open(args.domain_pairs))
        _sm.DOMAIN_PAIRS.clear()
        _sm.DOMAIN_PAIRS.update(custom)
        print(f"ℹ️   Domain pairs loaded from {args.domain_pairs} "
              f"({len(_sm.DOMAIN_PAIRS)} standard domains)\n")

    mapper.load_controls(controls_file)
    mapper.load_policies(policies_file, app=app_name)

    report = mapper.run(top_k=args.topk, threshold=args.threshold)
    mapper.print_report(report, verbose=args.verbose)

    # ── Save outputs ──────────────────────────────────────────────────────────
    mapper.save_report(report, f"{out_prefix}_report.json")

    try:
        mapper.save_csv(report, prefix=out_prefix)
    except Exception as e:
        print(f"⚠️  CSV export failed: {e}")

    try:
        import openpyxl
        mapper.save_xlsx(report, path=f"{out_prefix}_report.xlsx")
    except ImportError:
        print("⚠️  Excel export skipped — openpyxl not installed.")
        print("    Run:  pip install openpyxl")
    except Exception as e:
        print(f"⚠️  Excel export failed: {e}")

if __name__ == "__main__":
    main()
